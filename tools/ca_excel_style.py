# Shared Excel styling for CA Deployment workbook (document-like, no gridlines).
from copy import copy
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.marker import DataPoint as DP
from openpyxl.worksheet.page import PageMargins
from PIL import Image as PILImage
import os

NAVY = "1F4E79"
NAVY2 = "2E75B6"
GOLD = "C9A227"
RED = "C00000"
TEAL = "0D7377"
GREEN = "548235"
ORANGE = "C65911"
GRAY = "595959"
LIGHT = "F2F2F2"
PALE_BLUE = "D6EAF8"
PALE_GOLD = "FFF2CC"
PALE_GREEN = "E2EFDA"
PALE_ORANGE = "FCE4D6"
PALE_RED = "F8CBAD"
WHITE = "FFFFFF"
ROW_ALT = "EEF5FB"
EXAMPLE = "E2EFDA"
FORMULA = "DEEBF7"

thin = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
none_border = Border()

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(name="Calibri", size=11, bold=False, italic=False, color="000000"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def align(h="left", v="center", wrap=True, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def setup_sheet(ws, title, landscape=True, paper="A3"):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = True
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3 if paper == "A3" else ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.5, header=0.25, footer=0.25)
    ws.oddHeader.left.text = "Carrier Aggregation Deployment Guide"
    ws.oddHeader.right.text = "&A"
    ws.oddFooter.left.text = "Source: Huawei eRAN CA Feature Parameter Description, eRAN21.1 Issue 11 (2026-06-30)"
    ws.oddFooter.right.text = "Page &P of &N"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_properties.tabColor = NAVY
    ws.print_title_rows = "1:3"
    ws.freeze_panes = "A4"


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def put(ws, row, col, value, *, size=11, bold=False, italic=False, color="000000",
        fill_hex=None, h="left", v="center", wrap=True, border=False, indent=0):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font(size=size, bold=bold, italic=italic, color=color)
    cell.alignment = align(h, v, wrap, indent)
    if fill_hex:
        cell.fill = fill(fill_hex)
    if border:
        cell.border = thin
    return cell


def banner(ws, row, cols, text, fill_hex=NAVY, size=18, height=28):
    merge(ws, row, 1, row, cols)
    put(ws, row, 1, text, size=size, bold=True, color=WHITE, fill_hex=fill_hex, h="left", v="center")
    ws.row_dimensions[row].height = height
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).fill = fill(fill_hex)
        ws.cell(row=row, column=c).alignment = align("left", "center", True)
    return row + 1


def section(ws, row, cols, text, fill_hex=NAVY2, size=13):
    merge(ws, row, 1, row, cols)
    put(ws, row, 1, text, size=size, bold=True, color=WHITE, fill_hex=fill_hex, h="left", v="center")
    ws.row_dimensions[row].height = 22
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).fill = fill(fill_hex)
    return row + 1


def subsection(ws, row, cols, text, fill_hex=TEAL):
    merge(ws, row, 1, row, cols)
    put(ws, row, 1, text, size=12, bold=True, color=WHITE, fill_hex=fill_hex, h="left", v="center")
    ws.row_dimensions[row].height = 20
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).fill = fill(fill_hex)
    return row + 1


def note_bar(ws, row, cols, text, fill_hex=PALE_GOLD):
    merge(ws, row, 1, row, cols)
    put(ws, row, 1, text, size=10, italic=True, color="000000", fill_hex=fill_hex, h="left", v="center")
    ws.row_dimensions[row].height = max(28, 15 + 12 * (text.count("\n") + text.count(". ") // 8))
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).fill = fill(fill_hex)
    return row + 1


def body(ws, row, cols, text, height=None, fill_hex=None, size=11, bold=False):
    merge(ws, row, 1, row, cols)
    put(ws, row, 1, text, size=size, bold=bold, fill_hex=fill_hex, h="left", v="top")
    if height is None:
        n = max(1, text.count("\n") + 1)
        extra = len(text) // 140
        height = min(90, 16 + 13 * n + 4 * extra)
    ws.row_dimensions[row].height = height
    for c in range(2, cols + 1):
        if fill_hex:
            ws.cell(row=row, column=c).fill = fill(fill_hex)
        ws.cell(row=row, column=c).alignment = align("left", "top", True)
    return row + 1


def bullets(ws, row, cols, items, fill_hex=None):
    text = "\n".join(f"•  {it}" for it in items)
    return body(ws, row, cols, text, fill_hex=fill_hex)


def blank(ws, row, h=8):
    ws.row_dimensions[row].height = h
    return row + 1


def headers(ws, row, titles, fill_hex=NAVY, color=WHITE, height=22):
    for i, t in enumerate(titles, 1):
        put(ws, row, i, t, size=10, bold=True, color=color, fill_hex=fill_hex, h="center", v="center", border=True)
    ws.row_dimensions[row].height = height
    ws.auto_filter.ref = None  # keep document look; no autofilter chrome
    return row + 1


def table_row(ws, row, values, fills=None, bolds=None, height=None, center_cols=None):
    n = len(values)
    for i, v in enumerate(values, 1):
        fh = None
        if fills:
            fh = fills[i - 1] if i - 1 < len(fills) else fills[-1]
        b = False
        if bolds:
            b = bolds[i - 1] if i - 1 < len(bolds) else False
        h = "center" if center_cols and i in center_cols else "left"
        put(ws, row, i, v, size=10, bold=b, fill_hex=fh, h=h, v="top", border=True)
    if height is None:
        longest = max((len(str(v)) if v is not None else 0) for v in values)
        height = min(80, max(22, 16 + longest // 70 * 12))
    ws.row_dimensions[row].height = height
    return row + 1


def alt_fill(i, base=WHITE, alt=ROW_ALT):
    return alt if i % 2 else base


def insert_figure(ws, row, path, cols, caption, max_w=620, max_h=340):
    if not os.path.exists(path):
        return body(ws, row, cols, f"[Figure missing: {os.path.basename(path)}]", fill_hex=PALE_RED)
    im = PILImage.open(path)
    w, h = im.size
    scale = min(max_w / w, max_h / h, 1.0)
    nw, nh = int(w * scale), int(h * scale)
    img = XLImage(path)
    img.width = nw
    img.height = nh
    # place in column B
    ws.add_image(img, f"B{row}")
    # occupy several rows
    rows_needed = max(6, int(nh / 15) + 1)
    for i in range(rows_needed):
        ws.row_dimensions[row + i].height = 15
        merge(ws, row + i, 1, row + i, cols) if i == 0 else None
    # caption below
    cap_row = row + rows_needed
    merge(ws, cap_row, 1, cap_row, cols)
    put(ws, cap_row, 1, caption, size=9, italic=True, color=GRAY, h="center", v="center")
    ws.row_dimensions[cap_row].height = 18
    return cap_row + 1


def formula_box(ws, row, cols, title, formula, example, result):
    row = subsection(ws, row, cols, title, fill_hex=NAVY2)
    row = body(ws, row, cols, f"3GPP / Huawei condition:\n{formula}", fill_hex=FORMULA, size=11, bold=False)
    row = body(ws, row, cols, f"Worked example (sample values from the document or typical live-network settings):\n{example}\n\nResult: {result}", fill_hex=EXAMPLE)
    return row
