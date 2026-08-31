#!/usr/bin/env python3
"""Build Carrier Aggregation Deployment Excel (document-style, no gridlines).

Source: Huawei eRAN Carrier Aggregation Feature Parameter Description
        eRAN21.1 Issue 11 (2026-06-30)
"""
import os, sys
sys.path.insert(0, os.path.dirname(__file__))
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from ca_excel_style import *

FIG = "/workspace/ca_figures"
OUT = "/workspace/Carrier_Aggregation_Deployment.xlsx"
SRC = "Huawei eRAN Carrier Aggregation Feature Parameter Description, eRAN21.1 Issue 11 (2026-06-30)"
COLS = 10


def _fig(name):
    return os.path.join(FIG, name)


# ---------------------------------------------------------------------------
# COVER
# ---------------------------------------------------------------------------
def build_cover(wb):
    ws = wb.create_sheet("0. Cover & Index", 0)
    setup_sheet(ws, "Cover")
    set_widths(ws, [18, 22, 22, 22, 22, 22, 18, 18, 18, 18])
    r = 1
    r = banner(ws, r, COLS, "  LTE Carrier Aggregation (CA)  —  Deployment Workbook", size=22, height=36)
    r = body(ws, r, COLS, "Prepared from the attached Huawei feature parameter description. "
             "Layout is document-style (gridlines off). Values marked “Doc example” are taken "
             "verbatim from the MML samples in the source. Where the FPD does not publish a factory default, "
             "Default is shown as “See Parameter Reference”.", fill_hex=PALE_GOLD)
    r = blank(ws, r)
    r = section(ws, r, COLS, "Document identity")
    r = headers(ws, r, ["Item", "Value"] + [""] * 8)
    for i, (a, b) in enumerate([
        ("Source document", "eRAN  Carrier Aggregation Feature Parameter Description"),
        ("Product / version", "Huawei eRAN 21.1"),
        ("Issue / date", "Issue 11  /  2026-06-30"),
        ("Applicable RAT", "LTE FDD and LTE TDD"),
        ("Huawei feature family", "LAOFD / TDLAOFD / LEOFD / TDLEOFD / MRFD (see License sections)"),
        ("Purpose of this workbook", "Deployment-oriented extract: principles, triggers, leaving, parameters, MML, counters, KPI, licenses"),
        ("How to use", "Read sheets 1–2 for concepts. Deploy Step1 (2CC) first, then Step2–4. Each Step sheet is self-contained."),
    ]):
        r = table_row(ws, r, [a, b, "", "", "", "", "", "", "", ""], fills=[PALE_BLUE, WHITE] + [WHITE] * 8, bolds=[True] + [False] * 9, height=28)
        merge(ws, r - 1, 2, r - 1, COLS)
    r = blank(ws, r)
    r = section(ws, r, COLS, "Sheet map  (document sequence Ch.3 → Ch.8)")
    r = headers(ws, r, ["#", "Sheet", "Maps to document", "What you will find"] + [""] * 6)
    rows = [
        ("0", "0. Cover & Index", "—", "Identity, how to use, sheet map"),
        ("1", "1. Overview of CA", "Ch.3 Overview + Ch.4.1–4.3", "Definition, classification, protocol stack, events, scenarios"),
        ("2", "2. Type of CA Config", "Ch.4.4 Configuration Modes + Ch.11 Flexible CA + Ch.12", "Group-based, Adaptive, Flexible, Intelligent selection"),
        ("3", "3. Step1 2CC Config", "Ch.5 Downlink 2CC + Ch.4.6 carrier mgmt", "Full 2CC deployment package (template for later steps)"),
        ("4", "4. Step2 3CC Config", "Ch.6 Downlink 3CC", "Same structure as Step1; 3CC-specific switches / MML / KPI"),
        ("5", "5. Step3 4CC Config", "Ch.7 Downlink 4CC", "Same structure; 4CC switch + RLC timers"),
        ("6", "6. Step4 5CC Config", "Ch.8 Downlink 5CC", "Same structure; 5CC switch (license covered by 4CC+5CC)"),
    ]
    for i, rec in enumerate(rows):
        vals = list(rec) + [""] * 6
        r = table_row(ws, r, vals, fills=[alt_fill(i)] * 10, height=26)
        merge(ws, r - 1, 4, r - 1, COLS)
    r = blank(ws, r)
    r = section(ws, r, COLS, "Recommended live-network sequence")
    r = bullets(ws, r, COLS, [
        "Confirm overlapping coverage, neighbor relations (NoRmvFlag=FORBID_RMV_ENUM), licenses, UE Rel-10+ (2CC) / Rel-12+ (3CC+), EPC MBR.",
        "Choose Adaptive mode (recommended) unless the cluster is a small fixed co-coverage CA group.",
        "Activate Downlink 2CC, verify counters and KPI, then add 3CC → 4CC → 5CC by turning on CaDl3CCSwitch / CaDl4CCSwitch / CaDl5CCSwitch.",
        "Do not skip 2CC: 3CC/4CC/5CC list Downlink 2CC (and 3CC/4CC as applicable) as prerequisite functions.",
        "Huawei note: this FPD is activation guidance. Feature gains depend on the live scenario; contact Huawei service for optimization.",
    ], fill_hex=PALE_GREEN)
    ws.row_dimensions[r - 1].height = 90
    return ws


# ---------------------------------------------------------------------------
# OVERVIEW
# ---------------------------------------------------------------------------
def build_overview(wb):
    ws = wb.create_sheet("1. Overview of CA")
    setup_sheet(ws, "Overview")
    set_widths(ws, [16, 22, 18, 18, 16, 16, 16, 16, 16, 18])
    r = 1
    r = banner(ws, r, COLS, "  1.  Overview of Carrier Aggregation")
    r = note_bar(ws, r, COLS, "Document: Ch.3 Overview, Ch.4.1 Related Concepts, Ch.4.2 Protocol Stack, Ch.4.3 CA-related Events, Ch.4.5 Band Combinations.")

    r = section(ws, r, COLS, "1.1  Introduction — why CA exists")
    r = body(ws, r, COLS,
             "3GPP requires LTE-Advanced networks to provide a downlink peak data rate of 1 Gbit/s. "
             "Operators’ spectrum may not be contiguous, or may exceed the single-carrier bandwidth (max 20 MHz per CC). "
             "3GPP therefore introduced Carrier Aggregation (CA): aggregation of multiple contiguous or non-contiguous "
             "component carriers (CCs) into a larger bandwidth, to meet the 1 Gbit/s DL target and to raise UL/DL user peak rates.")
    r = insert_figure(ws, r, _fig("emb_p31_0.png"), COLS,
                      "Figure 3-1  CA — multiple component carriers aggregated for one UE  (source: Huawei FPD p.18)")

    r = section(ws, r, COLS, "1.2  How CA works (user-plane)")
    r = bullets(ws, r, COLS, [
        "PCell / PCC: the cell the CA UE camps on. Almost all CA switches are set on the PCell. The PCell cannot be deactivated or removed; it changes only at HO or RRC re-establishment.",
        "SCell / SCC: extra cell(s) configured by RRC Connection Reconfiguration from the PCell. An SCell may be DL-only or DL+UL. CCs are not actually aggregated until the SCell is activated by a MAC CE.",
        "Protocol stack: one PDCP and one RLC per radio bearer (RLC does not see CC count). MAC schedules each CC separately. Each CC has its own Uu transport channels and HARQ entities.",
        "Huawei product limit vs 3GPP: 3GPP allows up to 32 CCs × 20 MHz. Huawei CA currently supports DL max 8 CCs and UL max 2 CCs.",
        "Asymmetric CA is allowed: number of DL CCs ≥ number of UL CCs, and UL CCs must be a subset of the DL CC set.",
        "Each CC keeps the Release-8 frame structure (backward compatible). TDD cells that are aggregated must use the same UL/DL configuration (1 or 2).",
    ])
    ws.row_dimensions[r - 1].height = 110
    r = insert_figure(ws, r, _fig("emb_p38_0.png"), COLS,
                      "Figure 4-1  Protocol stack with CA enabled  (source: Huawei FPD p.25)")

    r = section(ws, r, COLS, "1.3  Classification")
    r = headers(ws, r, ["Dimension", "Type", "Meaning"] + [""] * 7)
    classif = [
        ("Service direction", "Downlink CA", "eNodeB aggregates intra- or inter-band carriers to widen DL bandwidth for a CA UE."),
        ("Service direction", "Uplink CA", "eNodeB aggregates intra- or inter-band carriers to widen UL bandwidth from a CA UE."),
        ("Number of carriers", "Fixed number (nCC)", "eNodeB selects a fixed number of CCs (2/3/4/5 DL, 2 UL)."),
        ("Number of carriers", "Non-fixed (Flexible CA)", "eNodeB picks the most suitable carriers from UE-reported combinations using coverage overlap, load, bandwidth, spectral efficiency. DL only."),
        ("RAT", "Single-RAT CA", "All FDD or all TDD."),
        ("RAT", "Multi-RAT CA", "FDD+TDD CA. Requires DL FDD+TDD CA or UL FDD+TDD CA to be enabled."),
        ("Site", "Intra-eNodeB CA", "Aggregated cells served by the same eNodeB. 3GPP TS 36.300 Annex J five scenarios."),
        ("Site", "Inter-eNodeB CA", "Cells served by different eNodeBs: Relaxed backhaul, or eNodeB coordination."),
    ]
    for i, rec in enumerate(classif):
        vals = list(rec) + [""] * 7
        r = table_row(ws, r, vals, fills=[alt_fill(i)] * 10, height=32)
        merge(ws, r - 1, 3, r - 1, COLS)

    r = blank(ws, r)
    r = section(ws, r, COLS, "1.4  Intra-eNodeB application scenarios  (3GPP TS 36.300 Annex J)")
    r = headers(ws, r, ["Scenario (Fig.)", "CA-group-based", "Adaptive (recommended)", "Operator note"] + [""] * 6)
    scen = [
        ("Co-coverage carriers (Fig. 3-2)", "Supported", "Supported", "Typical macro same-coverage dual-band."),
        ("Different-coverage carriers (Fig. 3-3)", "Supported", "Supported", "Hot-spot overlay."),
        ("Macro + edge coverage (Fig. 3-4)", "Supported (not recommended)", "Supported — recommended", "Adaptive is recommended in this scenario."),
        ("Macro + RRH (Fig. 3-5)", "FDD 1:1; TDD 1:N", "Supported 1:N (N≥1)", "Group-based FDD only when macro:RRH = 1:1."),
        ("Site + repeater (Fig. 3-6)", "Supported", "Supported", "Repeater extends one carrier."),
    ]
    for i, rec in enumerate(scen):
        vals = list(rec) + [""] * 6
        r = table_row(ws, r, vals, fills=[alt_fill(i)] * 10, height=26)
        merge(ws, r - 1, 4, r - 1, COLS)
    r = insert_figure(ws, r, _fig("emb_p33_0.png"), COLS, "Figure 3-2  Intra-eNodeB co-coverage carriers")
    r = insert_figure(ws, r, _fig("emb_p33_1.png"), COLS, "Figure 3-3  Intra-eNodeB different-coverage carriers")
    r = insert_figure(ws, r, _fig("emb_p34_0.png"), COLS, "Figure 3-4  Macro coverage + edge coverage")
    r = insert_figure(ws, r, _fig("emb_p34_1.png"), COLS, "Figure 3-5  Site + RRH")
    r = insert_figure(ws, r, _fig("emb_p34_2.png"), COLS, "Figure 3-6  Site + repeater")

    r = section(ws, r, COLS, "1.5  Inter-eNodeB scenarios")
    r = body(ws, r, COLS,
             "Huawei supports CA across eNodeBs in two ways. Related parameters of one of these scenarios must be set "
             "for DL or UL CA to take effect inter-eNodeB:\n"
             "• Relaxed backhaul (Ch.16) — DL 2–8 CC, UL 2 CC. FDD and TDD use different switches and different transport-quality requirements.\n"
             "• eNodeB coordination (Ch.17) — DL 2–8 CC, UL 2 CC. FDD supports centralized / distributed / hybrid; TDD supports centralized only.")

    r = section(ws, r, COLS, "1.6  CA function catalog in this eRAN version")
    r = headers(ws, r, ["Function", "UL", "DL", "CCs", "RAT", "Chapter"] + [""] * 4)
    funcs = [
        ("Downlink 2CC aggregation", "—", "Yes", "2", "FDD/TDD", "Ch.5  (this workbook Step1)"),
        ("Downlink 3CC aggregation", "—", "Yes", "3", "FDD/TDD", "Ch.6  (Step2)"),
        ("Downlink 4CC aggregation", "—", "Yes", "4", "FDD/TDD", "Ch.7  (Step3)"),
        ("Downlink 5CC aggregation", "—", "Yes", "5", "FDD/TDD", "Ch.8  (Step4)"),
        ("Downlink Massive CA", "—", "Yes", "6–8", "FDD only", "Ch.9  (not in Step1–4)"),
        ("Uplink 2CC aggregation", "Yes", "—", "2", "FDD/TDD", "Ch.10"),
        ("Flexible CA", "—", "Yes", "2–8", "FDD/TDD", "Ch.11  (see sheet 2)"),
        ("Intelligent serving-cell selection", "Yes", "Yes", "DL 2–5 / UL 2", "FDD/TDD", "Ch.12  (see sheet 2)"),
        ("Downlink FDD+TDD CA", "—", "Yes", "2–8", "FDD+TDD", "Ch.13"),
        ("Uplink FDD+TDD CA", "Yes", "—", "2", "FDD+TDD", "Ch.14"),
        ("Beamforming in SCells", "Yes", "Yes", "DL 2–5 / UL 2", "TDD only", "Ch.15"),
        ("Inter-eNB CA — relaxed backhaul", "Yes*", "Yes", "DL 2–8 / UL 2", "FDD/TDD", "Ch.16"),
        ("Inter-eNB CA — coordination", "Yes*", "Yes", "DL 2–8 / UL 2", "FDD/TDD", "Ch.17"),
        ("SCC buffer optimization", "—", "Yes", "2–8", "FDD", "Ch.18"),
    ]
    for i, rec in enumerate(funcs):
        vals = list(rec) + [""] * 4
        r = table_row(ws, r, vals, fills=[alt_fill(i)] * 10, height=22)
        merge(ws, r - 1, 6, r - 1, COLS)

    r = blank(ws, r)
    r = section(ws, r, COLS, "1.7  CA-related measurement events  (used for PCC anchoring and SCell management)")
    r = headers(ws, r, ["Event", "Meaning", "Entering condition (doc)", "Used for"] + [""] * 6)
    ev = [
        ("A1", "PCell quality exceeds a threshold", "Ms − Hys > Thresh   throughout TimeToTrig", "Start PCC anchoring (after UE camps on a non-highest-priority PCC)"),
        ("A2", "PCell / SCell quality drops below a threshold", "Ms + Hys < Thresh   throughout TimeToTrig", "SCell removal (CA A2). Also HO-related A2 can allow SCell config (RcvA2CfgSccSwitch)"),
        ("A3", "Neighbor better than PCell by an offset", "Mn+Ofn+Ocn−Hys > Ms+Ofs+Ocs+Off", "Connected-mode mobility (not CA-specific)"),
        ("A4", "Inter-frequency neighbor exceeds a threshold", "Mn+Ofn+Ocn−Hys > Thresh", "SCell add (measurement-based). Threshold = CarrAggrA4ThdRsrp + SCell/SCC A4 offset [+ extended Ofs]"),
        ("A5", "PCell below Thresh1 AND neighbor above Thresh2", "(Ms+Hys<Thresh1) AND (Mn+Ofn+Ocn−Hys>Thresh2)", "PCC anchoring inter-freq measurement. Thresh1 is always −43 dBm / −3 dB. Thresh2 = PCellA4RsrpThd or PccA4RsrpThd"),
        ("A6", "Intra-frequency neighbor of SCell better than SCell", "Mn+Ocn−Hys > Ms+Ocs+Off   (Hys always 1 dB)", "SCell change, PCell unchanged. Off = CaMgtCfg.CarrAggrA6Offset"),
    ]
    for i, rec in enumerate(ev):
        vals = list(rec) + [""] * 6
        r = table_row(ws, r, vals, fills=[alt_fill(i)] * 10, height=42)
        merge(ws, r - 1, 4, r - 1, COLS)
    r = note_bar(ws, r, COLS,
                 "Recommendation (Ch.4.3): select both CaA5HoEventSwitch and CaA5HoEventEnhSwitch on ENodeBAlgoSwitch so that event A4 is changed to A5 when SCells exist. "
                 "A4 measurement objects do not include serving cells, so without A5 the UE cannot be handed over from its PCell to its SCell. "
                 "If CaA6TimeToTrigger > 3 s (gap-assisted measurement timer), A6 reports cannot be sent.")

    r = section(ws, r, COLS, "1.8  Band combinations")
    r = bullets(ws, r, COLS, [
        "3GPP-defined combinations: with UMPT main control, combinations in 3GPP TS 36.101 V18.3.0 §5.6A.1.",
        "FDD 1.4/3 MHz cells can do FDD-only CA but cannot be PCell (only 6 RBs — PUCCH overload). TDD 1.4/3/5 MHz cannot do TDD-only CA.",
        "Private (operator-defined) combinations: PrivateCaBandComb + PrivateBand MOs. Private combination overrides 3GPP on conflict. Only one PrivateCaBandComb for 8CC (smallest PrivateCaCombId wins).",
        "UE reports CA capability in BandCombination IE. Network-requested signaling can collect up to 384 (128+256) combinations.",
    ])
    ws.row_dimensions[r - 1].height = 72
    r = note_bar(ws, r, COLS, "Peak rate is always capped by (1) BBP board peak of the PCell (example: LBBPd1 DL 450 Mbit/s) and (2) UE category (ue-CategoryDL / ue-CategoryUL, 3GPP TS 36.306).")
    return ws


# ---------------------------------------------------------------------------
# TYPE OF CA CONFIG
# ---------------------------------------------------------------------------
def build_types(wb):
    ws = wb.create_sheet("2. Type of CA Config")
    setup_sheet(ws, "Types")
    set_widths(ws, [18, 20, 20, 18, 16, 16, 16, 16, 16, 16])
    r = 1
    r = banner(ws, r, COLS, "  2.  Types of CA configuration  (how every type works)")
    r = note_bar(ws, r, COLS, "Document: Ch.4.4 Configuration Modes (group-based vs adaptive — adaptive is recommended), Ch.11 Flexible CA, Ch.12 Intelligent Selection of Serving Cell Combinations. "
                 "A mode change changes aggregatable cell combinations, license units, and inter-cell routes → SCell add/remove → counters L.Thrp.bits.DL.CAUser, L.Traffic.User.PCell/SCell.DL.Avg, L.Thrp.Time.DL.CAUser change.")

    # comparison
    r = section(ws, r, COLS, "2.1  Side-by-side comparison")
    r = headers(ws, r, ["Item", "CA-group-based", "Adaptive (recommended)", "Flexible CA", "Intelligent selection"] + [""] * 5)
    cmp_rows = [
        ("How enabled", "Deselect FreqCfgSwitch on ENodeBAlgoSwitch.CaAlgoSwitch", "Select FreqCfgSwitch AND AdpCaSwitch", "Group: license only. Adaptive: MultiCarrierFlexCaSwitch on CaMgtCfg.CellCaAlgoSwitch", "CaSmartSelectionSwitch on ENodeBAlgoSwitch.CaAlgoSwitch"),
        ("What you define", "CaGroup of cells (max 9 per group: FDD / TDD / FDDTDD)", "PccFreqCfg + SccFreqCfg frequencies", "Larger candidate SCC set; pick 1–7 SCCs (up to 8 inter-freq carriers)", "Best DL 2–5 / UL 2 combination from load, coverage, UE capability"),
        ("Who can aggregate", "Only cells inside the same CA group", "Only cells on the configured PCC/SCC frequencies", "From the candidate pool, prefer the combination with the largest number of CCs", "Selected combination, not a static pair"),
        ("Blind SCell", "SccBlindCfgSwitch + CaGroupSCellCfg.SCellBlindCfgFlag=TRUE", "CaGroupSCellCfg.SCellBlindCfgFlag=TRUE (one blind SCell per SCC frequency)", "Can combine with SmartCaFastSccCfgSwitch", "SmartCaFastSccCfgSwitch tries blind add during intelligent selection"),
        ("Route / capacity", "Static CaGroupSCellCfg. Max 1152 MOs / eNB", "Static + dynamic routes. 8 / 24 / 48 neighbors per cell depending on Dl2CCAckResShareSw and CaRouteNumberExtensionSwitch. PCC+SCC freqs ≤ 17.", "Same as parent mode", "If ON, max 9 PCC frequencies (else 16)"),
        ("License character", "Each cell in a ≥2-inter-freq group consumes LTE-A Introduction", "Each PCC-freq cell + participating SCC-freq cell consumes LTE-A Introduction", "Per cell when ≥3 inter-freq cells (group) or when switch ON (adaptive)", "Separate feature; see Ch.12"),
        ("When to use", "Small, fixed, co-coverage cluster; simple ops", "Multi-band / multi-site, overlapping coverage, growth. Huawei recommended.", "More than 2 candidate carriers; want max CC count", "Need load-aware serving-cell mix (DL and UL)"),
    ]
    for i, rec in enumerate(cmp_rows):
        vals = list(rec) + [""] * 5
        r = table_row(ws, r, vals, fills=[alt_fill(i)] * 10, height=52)
        merge(ws, r - 1, 5, r - 1, COLS)

    r = blank(ws, r)
    r = section(ws, r, COLS, "2.2  CA-group-based configuration — how it works")
    r = body(ws, r, COLS,
             "Cells are placed in a CaGroup. Only those cells can be aggregated. The eNodeB picks the highest PreferredPCellPriority cell as PCell "
             "(PCC anchoring) and the highest SCellPriority candidate as SCell. SCellPriority = 0 means the cell must never be an SCell. "
             "Blind add skips A4 and sends RRC Reconfiguration immediately.\n\n"
             "Rules: ≤9 cells per group; identical PreferredPCellPriority / SCellPriority / SpidGrpId recommended for intra-frequency cells in a group "
             "(eNodeB takes the highest if they differ). Candidate SCells must be in the same group as the PCell or the MO is wasted. "
             "If PCell is high-frequency (small coverage), blind-configure a low-frequency large-coverage SCell. If PCell is low-frequency (large coverage), "
             "do NOT blind-configure a high-frequency small cell — use A4; otherwise SCell scheduling efficiency collapses and the CA UE will not reach ~2× non-CA rate.")
    r = insert_figure(ws, r, _fig("emb_p44_0.png"), COLS, "Figure 4-2  CA-group-based configuration mode  (source: Huawei FPD p.30)")

    r = section(ws, r, COLS, "2.3  Adaptive configuration — how it works  (recommended)")
    r = body(ws, r, COLS,
             "You configure candidate PCC frequencies (PccFreqCfg) and, for each PCC, candidate SCC frequencies (SccFreqCfg). "
             "The eNodeB decides the actual PCell only after A5 reports (frequency-level, not cell-level, is configured). "
             "Blind add: only one SCellBlindCfgFlag=TRUE cell is allowed per SCC frequency; if several are flagged, which one wins is undefined and can change after reset.\n\n"
             "Limits: ≤16 PCC freqs (9 if CaSmartSelectionSwitch is ON). ≤16 SCC freqs per PCC (≤8 per operator unless SMART_CARRIER_SELECTION_SW or LOW_FREQ_SCS_OPT_SW). "
             "Total PCC+SCC frequencies ≤ 17. Max 1152 CaGroupSCellCfg, but CA route capacity / backplane may bind first.\n\n"
             "Dynamic routes age out after SCellAgingTime with no CA between the pair. Low-value routes can be penalized (CaRouteConfigPenaltyOfs ≠ 0). "
             "Doc recommendation in adaptive mode: configure one static route per cell to avoid wasting backplane; 48-neighbor mode needs UMPT+UBBP and raises CPU.")
    r = insert_figure(ws, r, _fig("emb_p45_0.png"), COLS, "Figure 4-3  Adaptive configuration mode  (source: Huawei FPD p.32)")

    r = section(ws, r, COLS, "2.4  Flexible CA — how it works")
    r = body(ws, r, COLS,
             "Flexible CA lets the eNodeB choose the most suitable DL carriers from a larger candidate set, using the UE-reported CA capability "
             "and carrier-management principles (not a fixed 2CC pair).\n"
             "• Group-based: no extra switch — apply license LAOFD-070201 / TDLAOFD-070201. In a group of ≥3 inter-frequency intra-duplex cells, each cell consumes one sales unit.\n"
             "• Adaptive: MultiCarrierFlexCaSwitch on CaMgtCfg.CellCaAlgoSwitch. ON → eNodeB can select 1–7 SCCs from up to 8 inter-frequency carriers; "
             "operating frequencies of the largest-CC combination take precedence. OFF → A4 measurement objects are limited to the number of SCCs the UE supports, "
             "ordered by SccPriority, so the largest combination may never be configured.\n"
             "Procedure: (1) UE reports CA capability after RRC setup / re-establishment / incoming HO. (2) eNodeB selects SCCs from capability + local deployment.")
    r = note_bar(ws, r, COLS, "License: FDD LAOFD-070201 Flexible CA from Multiple Carriers, model LT1SCAD2MC00, per cell. "
                 "TDD TDLAOFD-070201, model LT1SCADFMC00, per cell. License insufficiency can block cell activation.")

    r = section(ws, r, COLS, "2.5  Intelligent selection of serving cell combinations — how it works")
    r = body(ws, r, COLS,
             "Ch.12: the eNodeB selects DL combinations (2–5 CC) and UL combinations (2 CC) using triggers (traffic / measurement), "
             "UE capability, coverage overlap and load — rather than a static priority list alone. "
             "Requires CaSmartSelectionSwitch. When this switch is ON, PCC frequency count is capped at 9. "
             "SmartCaFastSccCfgSwitch (CaAlgoExtSwitch) can try blind SCell configuration to shorten time-to-CA; it requires CaSmartSelectionSwitch ON. "
             "Does not take effect together with some PCC-anchoring enhancements (e.g. UlCaCapbBasedPccSelectionSw).")

    r = section(ws, r, COLS, "2.6  Carrier-management state machine  (common to every type)")
    r = bullets(ws, r, COLS, [
        "PCC anchoring → SCell configuration → SCell change (A6) → SCell activation (MAC CE) → SCell deactivation (MAC CE) → SCell removal (RRC).",
        "Aggregation is live only after activation. Configured-but-deactivated SCells still consume C-RNTI / route resources.",
        "Enhancements: PccSmartCfgSwitch (load-based PCell) and SccSmartCfgSwitch (load-based SCell). Unknown load is treated as high load. High PCell load = PCell-served UEs > CellMaxPccNumber × PccUserNumberOffloadThd, OR BBP CPU > 70%.",
    ])
    ws.row_dimensions[r - 1].height = 56
    r = insert_figure(ws, r, _fig("emb_p91_0.png"), COLS, "Figure 4-6  SCell state transitions  (source: Huawei FPD p.78)")
    return ws
