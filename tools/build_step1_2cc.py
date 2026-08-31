# Step1 2CC sheet — full deployment package (template for 3/4/5CC).
import os, sys
sys.path.insert(0, os.path.dirname(__file__))
from ca_excel_style import *
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

FIG = "/workspace/ca_figures"
COLS = 10


def _fig(n):
    return os.path.join(FIG, n)


def add_param_header(ws, r):
    return headers(ws, r, ["SN", "MO Name", "Parameter name", "Default value", "Recommended value",
                           "Parameter description", "Notes"] + [""] * 3)


def add_param(ws, r, sn, mo, name, default, reco, desc, notes, related=False):
    fh = PALE_GOLD if related else alt_fill(sn)
    vals = [sn, mo, name, default, reco, desc, notes, "", "", ""]
    r = table_row(ws, r, vals, fills=[fh] * 10, bolds=[False, True, True, False, True, False, False],
                  center_cols={1}, height=min(78, 28 + len(desc) // 90 * 12))
    merge(ws, r - 1, 7, r - 1, COLS)
    return r


def add_mml_header(ws, r):
    return headers(ws, r, ["SN", "Seq", "Mode", "RAT", "Phase", "MML command (verbatim from document)", "Remarks"] + [""] * 3)


def add_mml(ws, r, sn, seq, mode, rat, phase, cmd, remark):
    vals = [sn, seq, mode, rat, phase, cmd, remark, "", "", ""]
    r = table_row(ws, r, vals, fills=[alt_fill(sn)] * 10, bolds=[False, False, False, False, False, True, False],
                  center_cols={1, 2}, height=min(70, 24 + len(cmd) // 80 * 12))
    merge(ws, r - 1, 7, r - 1, COLS)
    return r


def add_ctr_header(ws, r):
    return headers(ws, r, ["SN", "Counter ID", "Counter Name", "Function / what it measures", "Use"] + [""] * 5)


def add_ctr(ws, r, sn, cid, name, fn, use):
    vals = [sn, cid, name, fn, use, "", "", "", "", ""]
    r = table_row(ws, r, vals, fills=[alt_fill(sn)] * 10, bolds=[False, False, True, False, False],
                  center_cols={1}, height=32)
    merge(ws, r - 1, 5, r - 1, COLS)
    return r


def peak_chart(ws, r, title, cats, fdd, tdd, anchor):
    """Write a small data block then a bar chart of theoretical peaks."""
    r = subsection(ws, r, COLS, title)
    start = r
    ws.cell(row=r, column=1, value="MIMO / MCS")
    ws.cell(row=r, column=2, value="FDD (Mbit/s)")
    ws.cell(row=r, column=3, value="TDD (Mbit/s)")
    for c in range(1, 4):
        ws.cell(row=r, column=c).font = font(size=10, bold=True, color=WHITE)
        ws.cell(row=r, column=c).fill = fill(NAVY)
        ws.cell(row=r, column=c).alignment = align("center", "center")
    r += 1
    for i, (c, a, b) in enumerate(zip(cats, fdd, tdd)):
        ws.cell(row=r, column=1, value=c)
        ws.cell(row=r, column=2, value=a)
        ws.cell(row=r, column=3, value=b)
        for col in range(1, 4):
            ws.cell(row=r, column=col).fill = fill(alt_fill(i))
            ws.cell(row=r, column=col).font = font(size=10)
            ws.cell(row=r, column=col).alignment = align("center", "center")
            ws.cell(row=r, column=col).border = thin
        r += 1
    end = r - 1
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Theoretical DL peak (20 MHz/CC; TDD UL/DL cfg 2, SS 7)"
    chart.y_axis.title = "Mbit/s"
    chart.style = 10
    data = Reference(ws, min_col=2, min_row=start, max_col=3, max_row=end)
    cats_ref = Reference(ws, min_col=1, min_row=start + 1, max_row=end)
    chart.add_data(data, from_rows=False, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    chart.legend.position = "b"
    chart.width = 18
    chart.height = 8
    ws.add_chart(chart, f"E{start}")
    # occupy rows for chart
    for i in range(14):
        if ws.row_dimensions[start + i].height is None or ws.row_dimensions[start + i].height < 18:
            ws.row_dimensions[start + i].height = 18
    return max(r, start + 15)


def trigger_leave_block(ws, r, cc_label="2CC"):
    """Shared Ch.4.6 triggering / leaving content with charts + numerical examples."""
    r = section(ws, r, COLS, f"B.  Triggering conditions  ({cc_label} uses the common CA carrier-management procedure in Ch.4.6)")
    r = note_bar(ws, r, COLS, "Charts below are the original Huawei figures. Sample calculations use document-fixed constants "
                 "(A5 Thresh1 RSRP = −43 dBm, RSRQ = −3 dB; A6 Hys = 1 dB; A5 measurement timer = 3 s) "
                 "plus the MML example thresholds PCellA4RsrpThd / PccA4RsrpThd = −105 dBm, offsets = 0.")

    r = subsection(ws, r, COLS, "B1.  PCC anchoring — when it starts")
    r = bullets(ws, r, COLS, [
        "Enable EnhancedPccAnchorSwitch (recommended). If both PccAnchorSwitch and EnhancedPccAnchorSwitch are ON, only Enhanced takes effect.",
        "eNodeB delivers A1 measurement to the CA UE when: (1) the access / re-establishment / HO target cell is NOT the highest PCell/PCC priority, AND (2) the UE meets attribute conditions.",
        "UE attribute conditions: not eMBMS; not emergency call; not high-speed mobility state; if PCC_ANCHOR_HO_FORBID_SW is ON, speed ≤ 30 km/h; not eMTC; QCI1/VoLTE and QCI65/66 MCPTT rules (see 19.7 VoLTE / McpttVoiceCaSwitch).",
        "A1 RSRP threshold = CaMgtCfg.EnhancedPccAnchorA1ThdRsrp. SCell config is allowed before A1 is reported.",
        "On A1: if HoWithSccCfgSwitch OFF → remove all SCells then PCC-anchor; if ON → PCC-anchor keeping SCells.",
        "In the current serving cell, A1 is reported only once. CaTrafficTriggerSwitch ON allows PCC anchoring on unnecessary incoming HO (risk of ping-pong); OFF = necessary HO only (recommended to avoid ping-pong).",
    ])
    ws.row_dimensions[r - 1].height = 110
    r = formula_box(ws, r, COLS,
                    "Event A1  —  entering condition (3GPP / Ch.4.3)  +  sample",
                    "A1 enters when  (Ms − Hys > Thresh)  is true throughout TimeToTrig.\nMs = PCell RSRP (dBm).  Thresh = CaMgtCfg.EnhancedPccAnchorA1ThdRsrp.",
                    "Assume: Ms = −82 dBm (UE in good PCell coverage), Hys = 1 dB, Thresh = −90 dBm, TimeToTrig met.\n"
                    "Compute: Ms − Hys = −82 − 1 = −83 dBm.  Compare to Thresh −90 dBm:  −83 > −90  → TRUE.",
                    "A1 IS triggered. eNodeB starts PCC anchoring (group Fig. 4-4 or adaptive Fig. 4-5). "
                    "If Ms = −95 dBm: −95 − 1 = −96 which is NOT > −90 → A1 does not trigger, UE stays on current PCell.")
    r = insert_figure(ws, r, _fig("emb_p85_0.png"), COLS,
                      "Figure 4-4  CA-group-based PCC anchoring procedure  (source: Huawei FPD p.72)  —  sample values in figure notes: PCellA4RsrpThd=−105 dBm, A5 Thresh1=−43 dBm, timer=3 s")
    r = body(ws, r, COLS,
             "Group-based sample (Doc MML): Cell0 PreferredPCellPriority=2, Cell1 PreferredPCellPriority=1, both PCellA4RsrpThd=−105, PCellA4RsrqThd=−20.\n"
             "UE accesses Cell1 (priority 1). Highest priority in group is Cell0 (priority 2) → Cell0 is candidate. UE capable on Cell0 frequency → eNodeB sends A5 "
             "(Thresh1 RSRP always −43 dBm / RSRQ −3 dB; Thresh2 RSRP = −105 dBm from PCellA4RsrpThd; InterFreqHoA4TrigQuan = RSRP or RSRQ or BOTH). Timer 3 s.\n"
             "If neighbor Cell0 RSRP Mn = −98 dBm, Ofn=Ocn=0, Hys=1:  Mn+Ofn+Ocn−Hys = −99 dBm.  −99 > −105  AND  PCell quality below −43 (always true at realistic RSRP) → A5 fires → HO to Cell0 as new PCell.\n"
             "If Mn = −112 dBm: −113 is NOT > −105 → no A5 within 3 s → try next candidate; if none, Cell1 remains PCell.",
             fill_hex=EXAMPLE)
    ws.row_dimensions[r - 1].height = 88
    r = insert_figure(ws, r, _fig("emb_p87_0.png"), COLS,
                      "Figure 4-5  Adaptive PCC anchoring procedure  (source: Huawei FPD p.74)  —  PccFreqCfg.PccA4RsrpThd recommended > InterFreqHoGroup.InterFreqHoA4ThdRsrp")
    r = note_bar(ws, r, COLS, "Adaptive difference: only frequency-level PCC is configured, so the PCell cell is chosen from the first A5 report. "
                 "If no SccFreqCfg exists for a PccFreqCfg, A5 for that PCC is not delivered. Recommendation: PccA4RsrpThd > coverage HO A4 threshold so PCC anchoring does not steal coverage HOs.")

    r = subsection(ws, r, COLS, "B2.  SCell configuration — when it starts")
    r = bullets(ws, r, COLS, [
        "Starts after RRC setup in: initial access, incoming RRC re-establishment, or incoming HO — AND SRB2 + default DRB exist — AND SCell count < max — AND not emergency call — AND not CSFB.",
        "CaTrafficTriggerSwitch ON: SCell config only when traffic also meets SCell activation conditions (blind add does not take effect). OFF: triggered by those traffic conditions OR by RRC setup.",
        "PCell-served UE cap: if number of PCell-served UEs ≥ CellMaxPccNumber, SCell config is prohibited. Resume only when the number falls below 90% of CellMaxPccNumber. SCell-served UEs are not counted. Lowering CellMaxPccNumber does not remove already-configured SCells.",
        "CPU: eNodeBFlowCtrlPara.CaFlowCtrlThld. Below (thld−10): no impact. At thld: 80% probability; then 60/40/20/10%; at thld+5 to 100%: no SCell config. BBP CPU [0,70): no impact; [70,90): no SCell config for new access/HO/re-est; [90,100]: no SCell config for any CA UE.",
        "UE-number control: Forbid2/3/4/5ccUeRatioThld. Doc example for large UE count: 90 / 80 / 70 / 60. Default 100 = function OFF. Constraint: Forbid2 > Forbid3 > Forbid4 > Forbid5 when not 100. Exit hysteresis = threshold − 5% (floor 1%).",
        "FREQ_MEAS_FLAG of EutranInterNFreq.AggregationAttribute must stay selected or the frequency is not a candidate.",
        "For nCC, the eNodeB prefers the maximum supported CC count. If fewer SCells succeed, it retries every SccCfgInterval while activation traffic conditions remain true.",
    ])
    ws.row_dimensions[r - 1].height = 130
    r = insert_figure(ws, r, _fig("emb_p93_0.png"), COLS,
                      "Figure 4-7  Control of the maximum permissible number of PCell-served UEs over SCell configuration  (source: Huawei FPD p.79)")
    r = formula_box(ws, r, COLS,
                    "Event A4  —  SCell add (measurement-based)  +  sample",
                    "A4 enters when  (Mn + Ofn + Ocn − Hys > Thresh)  throughout CaMgtCfg.CaA4TimeToTrigger.\n"
                    "Effective Thresh (group, initial/HO/re-est) = CaMgtCfg.CarrAggrA4ThdRsrp + CaGroupSCellCfg.SCellA4Offset + SccFreqCfg.SccA2RsrpThldExtendedOfs  (all on PCell).\n"
                    "Effective Thresh (adaptive, initial/HO/re-est) = CarrAggrA4ThdRsrp + SccFreqCfg.SccA4Offset + SccA2RsrpThldExtendedOfs.\n"
                    "If SCell config is periodically triggered by traffic (not initial), the extended Ofs is NOT added (adaptive).",
                    "Doc MML: SCellA4Offset=0, SccA4Offset=0. Assume CarrAggrA4ThdRsrp = −105 dBm (same order as PCell A4 example), Hys=1 dB, Ofn=Ocn=0.\n"
                    "Thresh = −105 + 0 + 0 = −105 dBm.\n"
                    "Candidate SCell RSRP Mn = −100 dBm → Mn−Hys = −101.  −101 > −105 → A4 IS triggered → SCell add.\n"
                    "If Mn = −110 dBm → −111 which is NOT > −105 → A4 not triggered (no measurement-based add).\n"
                    "Raising SCellA4Offset by +6 dB moves Thresh to −99 dBm: then Mn=−100 gives −101 which is NOT > −99 → SCell add becomes harder (doc: greater offset = lower probability of configuring an SCell).",
                    "Use A4 for high-frequency small SCells under a low-frequency PCell. Use blind add (skip A4) when the SCell is low-frequency large coverage under a high-frequency PCell — A4 would always fire.")
    r = insert_figure(ws, r, _fig("emb_p99_0.png"), COLS,
                      "Figure 4-8  CA-group-based SCell configuration procedure for UL/DL 2CC  (source: Huawei FPD p.86). For nCC there are (n−1) candidate SCells in a selection.")
    r = insert_figure(ws, r, _fig("emb_p103_0.png"), COLS,
                      "Figure 4-9  Adaptive SCell configuration procedure for UL/DL 2CC  (source: Huawei FPD p.90)")

    r = subsection(ws, r, COLS, "B3.  SCell activation — when aggregation actually starts")
    r = insert_figure(ws, r, _fig("emb_p110_0.png"), COLS, "Figure 4-11  SCell activation  (source: Huawei FPD p.97)")
    r = formula_box(ws, r, COLS,
                    "Traffic-volume activation  (Table 4-14)  +  sample",
                    "DL SCell activates when BOTH:\n"
                    "  (1) RLC buffered data > max(Uu RLC rate × BufferDelayThd, BufferLenThd)\n"
                    "  (2) Delay of first RLC PDU > BufferDelayThd\n"
                    "Non-NSA: BufferDelayThd = CaMgtCfg.ActiveBufferDelayThd ; BufferLenThd = CaMgtCfg.ActiveBufferLenThd.\n"
                    "NSA: NsaDcLteScellActBfrDelThld / NsaDcLteScellActBfrLenThld.\n"
                    "UL extra: BSR > BufferLenThd throughout UlCaActiveTimeToTrigger, and TTI bundling not in effect.\n"
                    "CaInstantlyJudgeSwitch ON = millisecond instantaneous traffic; OFF = filtered second-level (better for large-file DL).\n"
                    "MAC CE in subframe n → SCell active in subframe n+x with x=8 (FDD) or x≥8 (TDD).",
                    "Assume non-NSA, ActiveBufferLenThd = 10000 bytes, ActiveBufferDelayThd = 50 ms, Uu RLC rate = 2 Mbit/s = 250000 bytes/s.\n"
                    "max(rate × delay, len) = max(250000×0.050, 10000) = max(12500, 10000) = 12500 bytes.\n"
                    "If RLC buffer = 20000 bytes AND first-PDU delay = 80 ms (> 50 ms) → BOTH true → eNodeB sends MAC CE to activate SCell.\n"
                    "If ActiveBufferLenThd = 0 (explicitly not recommended): eNodeB activates as soon as it intends to send any data, then often deactivates immediately → BBP flow control risk.",
                    "SCell becomes usable only after activation. Heavy-load block: if ScellNoActivationUeNumThld ≠ 0 and UE count (non-CA + PCell CA + activated-SCell CA) exceeds it, no new activation quota. Recommended ON for FDD 1.4/3/5 MHz and TDD 5 MHz. Not compatible with LBBPc.")

    r = section(ws, r, COLS, f"C.  Leaving conditions  (SCell deactivation + SCell removal)")
    r = insert_figure(ws, r, _fig("emb_p113_0.png"), COLS, "Figure 4-12  SCell deactivation triggers  (source: Huawei FPD p.100)")

    r = subsection(ws, r, COLS, "C1.  Traffic-volume deactivation  (needs CaMgtCfg.CarrierMgtSwitch = ON on PCell)")
    r = formula_box(ws, r, COLS,
                    "Table 4-15  +  sample  (recommend DeactiveBufferLenThd < ActiveBufferLenThd)",
                    "DL deactivation when EVERY E-RAB meets BOTH:\n"
                    "  Uu RLC rate ≤ CaMgtCfg.DeactiveThroughputThd\n"
                    "  RLC buffer ≤ CaMgtCfg.DeactiveBufferLenThd\n"
                    "Then eNodeB sends MAC CE to deactivate SCells.\n"
                    "UL SCell: deactivate only if UL AND DL both meet (UL: BSR ≤ DeactiveBufferLenThd AND UL Uu rate ≤ DeactiveThroughputThd). "
                    "If only DL meets: deactivate DL-only SCells.\n"
                    "If CaTrafficTriggerSwitch ON: RRC Reconfiguration removes SCells immediately after traffic-based deactivation "
                    "(except NSA CA UEs when ScgAdditionBufferLenThld = 0).",
                    "Assume DeactiveThroughputThd = 256 kbit/s, DeactiveBufferLenThd = 2000 bytes, ActiveBufferLenThd = 10000 bytes (hysteresis).\n"
                    "UE browsing idle: Uu rate = 80 kbit/s (≤ 256) and RLC buffer = 400 bytes (≤ 2000) on every E-RAB → DEACTIVATE.\n"
                    "UE still downloading: Uu rate = 8 Mbit/s → condition fails → SCell stays active.\n"
                    "If DeactiveBufferLenThd were 15000 (> Active 10000), ping-pong activate/deactivate would occur — this is why the document requires deactive length < active length.",
                    "MAC CE deactivates; with CaTrafficTriggerSwitch ON the SCell is also removed (RRC).")

    r = subsection(ws, r, COLS, "C2.  Channel-quality deactivation  (CarrierMgtSwitch = ON; recommended when blind SCell is used or A2 is OFF)")
    r = body(ws, r, COLS,
             "Does NOT run if CaSccSuspendSwitch is ON (SCell suspends scheduling instead) or if SCC_SPCT_EFF_STOP_SCH_SW is ON.\n"
             "Non-relaxed BH: deactivate when spectral efficiency of reported SCell CQI < Rel-8 SE of CaMgtCfg.SccDeactCqiThd (single-codeword). "
             "Value 5 is recommended. Value 0 disables CQI-based deactivation.\n"
             "Relaxed BH: use RelaxedBHSccDeactCqiThd instead.\n"
             "After CQI-based deactivation the UE stops reporting that SCell CQI. Reactivation by traffic/voice is gated by SccReactivationTime (PCell). "
             "SCell is NOT immediately removed, even if CaTrafficTriggerSwitch is ON.\n"
             "Sample: SccDeactCqiThd = 5. UE reports CQI 3 (SE below CQI5 SE) → MAC CE deactivate. UE reports CQI 8 → keep SCell.",
             fill_hex=EXAMPLE)
    ws.row_dimensions[r - 1].height = 88

    r = subsection(ws, r, COLS, "C3.  Residual BLER deactivation")
    r = body(ws, r, COLS,
             "DL: 10 consecutive residual BLER events. If SccQuietTime = 0 → immediate MAC CE. If SccQuietTime ≠ 0 → stop scheduling, send 1-byte probe; ACK resumes; NACK decreases CQI by basic step (+ SccDetectCqiDecreaseStep if ≠ 0); after max probe failures → deactivate.\n"
             "UL: SccQuietTime must be 0. 40 consecutive UL residual BLER → MAC CE. If SccDeactByUlDtxSwitch ON, also deactivate on activation-status mismatch. Do not enable SccDeactByUlDtxSwitch in high UL load (SCell rate drops). Enable it on every eNodeB in the CA set (PCell and SCells).")

    r = subsection(ws, r, COLS, "C4.  RLC retransmission-count deactivation")
    r = body(ws, r, COLS,
             "Relaxed BH: if ScellDeactRlcRetransNumThd = 0 on PCell, deactivate all SCells when total RLC retx on PCell+SCells > 20. If Thd ≠ 0, that threshold is used.\n"
             "Purpose: high SCell RBLER drives MAC retx → RLC retx → service drop. Deactivate SCell to protect drop rate.")

    r = subsection(ws, r, COLS, "C5.  SCell removal (RRC)  —  Event A2")
    r = insert_figure(ws, r, _fig("emb_p120_0.png"), COLS, "Figure 4-13  SCell removal  (source: Huawei FPD p.106)")
    r = formula_box(ws, r, COLS,
                    "Event A2  —  SCell leave / removal  +  sample",
                    "A2 enters when  (Ms + Hys < Thresh)  throughout CaMgtCfg.CaA2TimeToTrigger.\n"
                    "Effective Thresh (group) = CaMgtCfg.CarrAggrA2ThdRsrp + CaGroupSCellCfg.SCellA2Offset + SccFreqCfg.SccA2RsrpThldExtendedOfs  (PCell side).\n"
                    "Adaptive: CarrAggrA2ThdRsrp + SccFreqCfg.SccA2Offset  (clamped to [−140, −43] dBm).\n"
                    "Constraint: CarrAggrA4ThdRsrp must be greater than CarrAggrA2ThdRsrp (add threshold stricter than leave).\n"
                    "SccA2RmvSwitch must be selected to deliver A2 for SCells that were added by A4.\n"
                    "Greater SCellA2Offset / SccA2Offset → higher A2 threshold → SCell removed more easily.",
                    "Assume CarrAggrA2ThdRsrp = −115 dBm, SCellA2Offset = 0, Hys = 1 dB (typical leave hysteresis vs A4 −105).\n"
                    "Thresh = −115 dBm. SCell RSRP Ms = −118 dBm → Ms+Hys = −117.  −117 < −115 → A2 IS triggered → SCell removal.\n"
                    "If Ms = −100 dBm → −99 which is NOT < −115 → keep SCell.\n"
                    "If SCellA2Offset = +6 → Thresh = −109. Then Ms = −112 → −111 < −109 → leave fires earlier (UE still in decent coverage). "
                    "Doc: greater A2 offset = higher probability of removing an SCell.",
                    "A2 is the primary radio leaving condition. Traffic/CQI/BLER/RLC are MAC deactivation (C1–C4); A2 is RRC removal.")

    r = subsection(ws, r, COLS, "C6.  SCell change (not leave of CA — replace SCell, keep PCell)  —  Event A6")
    r = insert_figure(ws, r, _fig("emb_p109_0.png"), COLS, "Figure 4-10  SCell change  (source: Huawei FPD p.96)")
    r = formula_box(ws, r, COLS,
                    "Event A6  +  sample  (needs SccModA6Switch)",
                    "A6 enters when  (Mn + Ocn − Hys > Ms + Ocs + Off)  throughout CaMgtCfg.CaA6TimeToTrigger.\n"
                    "Hys is always 1 dB. Off = CaMgtCfg.CarrAggrA6Offset (PCell). Ocn/Ocs = EutranInterFreqNCell.CellIndividualOffset.\n"
                    "If CaA6TimeToTrigger > 3 s (gap timer), A6 cannot be reported.",
                    "Assume Off = 3 dB, Ocn=Ocs=0, Hys=1, current SCell Ms = −95 dBm, intra-freq neighbor Mn = −88 dBm.\n"
                    "Left: Mn+Ocn−Hys = −88 − 1 = −89.  Right: Ms+Ocs+Off = −95 + 3 = −92.  −89 > −92 → A6 IS triggered.\n"
                    "eNodeB ranks reported cells by RSRP and changes SCell to the best cell that can set up a data link.\n"
                    "If Off = 10 dB, right side = −85; −89 is NOT > −85 → no change (doc: greater CarrAggrA6Offset = lower probability of SCell change).",
                    "CA continues; only the SCell identity changes.")
    return r


def build_2cc(wb):
    ws = wb.create_sheet("3. Step1 2CC Config")
    setup_sheet(ws, "2CC")
    set_widths(ws, [8, 28, 32, 22, 28, 55, 42, 12, 12, 12])
    r = 1
    r = banner(ws, r, COLS, "  Step 1 —  Downlink 2CC configuration   (document Ch.5 + common Ch.4.6)")
    r = note_bar(ws, r, COLS, "This sheet is the master template. Steps 2–4 reuse the same section order. "
                 "2CC is a prerequisite of 3CC/4CC/5CC. Deploy and verify 2CC before turning on CaDl3CCSwitch / CaDl4CCSwitch / CaDl5CCSwitch.")

    # A Principal
    r = section(ws, r, COLS, "A.  Principal  (Ch.5.1–5.2)")
    r = insert_figure(ws, r, _fig("emb_p125_0.png"), COLS, "Figure 5-1  Downlink 2CC aggregation  (source: Huawei FPD p.112)")
    r = body(ws, r, COLS,
             "Downlink 2CC aggregates two intra- or inter-band carriers to raise DL bandwidth. It works intra-eNodeB, inter-eNodeB (coordination), and inter-eNodeB (relaxed backhaul).\n"
             "Switch map (Table 5-1):\n"
             "• Group-based FDD: any BW ≤ 40 MHz — CaDl2CCExtSwitch not required.\n"
             "• Group-based TDD: any BW ≤ 40 MHz — CaDl2CCExtSwitch not required.\n"
             "• Adaptive FDD: BW ≤ 20 MHz — no extra switch;  20 MHz < BW ≤ 40 MHz — select CaDl2CCExtSwitch on each possible PCell/SCell pair.\n"
             "• Adaptive TDD: BW ≤ 30 MHz — no extra switch;  30 MHz < BW ≤ 40 MHz — select CaDl2CCExtSwitch.\n"
             "FDD vs TDD difference (Ch.2.4): maximum aggregated bandwidth 20 MHz (FDD basic) vs 30 MHz (TDD basic) before the 40 MHz license/switch.")
    ws.row_dimensions[r - 1].height = 96
    r = peak_chart(ws, r, "Theoretical peak data rates for DL 2CC  (Table 5-2, unit Mbit/s)",
                   ["2x2 + 64QAM", "2x2 + 256QAM", "4x4 + 64QAM", "4x4 + 256QAM"],
                   [299.6, 391.6, 599.7, 783.3],
                   [224, 286, 434, 572], "2cc")
    r = body(ws, r, COLS,
             "Caps: PCell BBP peak (e.g. LBBPd1 DL 450 Mbit/s) and UE ue-CategoryDL. "
             "Network impact highlights: one CA UE = 1 RRC license but 1 HW unit per serving cell (nCC → admitted UEs → 1/n). "
             "PRB usage of the network usually rises (burst traffic, load balance). PUCCH/PUSCH/PDCCH overhead on PCell rises. "
             "Periodic CQI of CA UEs can drop (CSI discarded when colliding with ACK/NACK on PUCCH 1b). "
             "Select SriPeriodCfgOptSW. Select Dl2CCAckResShareSw only knowing it raises CA UE count but can lower per-UE rate.")

    r = trigger_leave_block(ws, r, "2CC")

    # D prereq + impacts
    r = section(ws, r, COLS, "D.  Prerequisite, mutually exclusive, and mutually impacted features   (one feature per row; all related parameters listed)")
    r = note_bar(ws, r, COLS, "Ch.5.3.3.1 Prerequisite = None. Ch.5.3.3.2 Mutually exclusive must be OFF. Ch.5.3.3.3 Function impacts (RAN category excerpt — the FPD table runs pp.119–156). "
                 "Relation = Exclusive (cannot coexist) / Impact (can coexist, behaviour changes) / Prerequisite.")
    r = headers(ws, r, ["SN", "Relation", "RAT", "Feature", "Parameter(s)  —  one row one feature", "Required action", "Impact / note"] + [""] * 3)
    impacts = [
        ("1", "Prerequisite", "FDD/TDD", "—", "None", "No other feature must be ON before 2CC", "Ch.5.3.3.1"),
        ("2", "Exclusive", "FDD", "Super combined cell", "CellAlgoSwitch.SfnAlgoSwitch → SuperCombCellSwitch  must be OFF", "Deactivate super combined cell", "Ch.5.3.3.2"),
        ("3", "Exclusive", "FDD", "Cell radius > 100 km", "Extended Cell Range (no dedicated CA switch)", "Do not use 2CC in >100 km cells", "Ch.5.3.3.2"),
        ("4", "Exclusive", "FDD/TDD", "In-band relay", "CellAlgoSwitch.RelaySwitch → InBandRelayDeNbSwitch; ReBTS InBandRelayReNbSwitch", "Deactivate in-band relay", "Ch.5.3.3.2"),
        ("5", "Exclusive", "TDD", "Enhanced symbol power saving", "CellAlgoSwitch.DlSchSwitch → MBSFNShutDownSwitch  must be OFF", "Deactivate MBSFN shutdown energy saving", "Ch.5.3.3.2"),
        ("6", "Exclusive", "FDD/TDD", "Intelligent multi-RF-module coordinated energy saving", "EnodebPwrSavingAlgo.PwrSavingAlgoSwitch → MIE_BASED_EE_CARR_SEL_SW", "Cannot enable together with CA-group-based mode", "Footnote a in Ch.5.3.3.2"),
        ("7", "Impact", "FDD/TDD", "User-number-based connected-mode MLB", "CellAlgoSwitch.MlbAlgoSwitch → InterFreqMlbSwitch; CellMLB.MlbTriggerMode=UE_NUMBER_ONLY; CellMLB.InterFreqUeTrsfType → SynchronizedUE; CellAlgoSwitch.EnhancedMlbAlgoSwitch → CaUserLoadTransferSw", "If CaUserLoadTransferSw OFF, eNodeB filters CA UEs (PCell or SCell) out of MLB UE selection", "Related: CaUserLoadTransferSw sits with MLB, not CaAlgoSwitch — listed here because it gates CA UE transfer"),
        ("8", "Impact", "FDD/TDD", "NSA/SA selection based on DL traffic volume", "EnodebAlgoExtSwitch.MultiNetworkingOptionOptSw → LTE_FDD_NSA_SA_DL_SEL_OPT_SW or LTE_TDD_NSA_SA_DL_SEL_OPT_SW", "EN-DC combination frozen until SCG release → eNodeB cannot add SCells from DL CA selection", "Handover Request selectedbandCombinationInfoEN-DC-v1540"),
        ("9", "Impact", "FDD", "NSA/SA selection based on UL coverage", "EnodebAlgoExtSwitch.MultiNetworkingOptionOptSw → LTE_FDD_NSA_SA_UL_SEL_OPT_SW", "Same freeze of EN-DC combination; SCells cannot follow DL CA selection", "Related to row 8 (same MO, different bit)"),
        ("10", "Impact", "FDD/TDD", "CSPC (centralized Cloud BB)", "CspcAlgoPara.CspcAlgoSwitch (FDD) / CspcAlgoPara.TddCspcAlgoSwitch (TDD)", "CA UE scheduling priority uses sum of rates on all CCs → CA UEs scheduled less often vs non-CA", "Consider differentiated scheduling"),
        ("11", "Impact", "FDD/TDD", "Zero Guard Band Between Contiguous Intra-Band Carriers", "ContigIntraBandCarr.ContigIntraBandCarrSw → CONTIG_INTRA_BAND_CARR_SW; ENodeBAlgoSwitch.CaAlgoExtSwitch → ContigIntraBandCaSwitch", "Contiguous intra-band CA does not take effect even if ContigIntraBandCaSwitch is ON. Intra-band non-contiguous CA still can in overlap scenarios", "Related: ContigIntraBandCaSwitch listed immediately because it is the CA-side companion"),
        ("12", "Impact", "FDD", "Short TTI", "CellShortTtiAlgo.SttiAlgoSwitch → SHORT_TTI_SW", "CA does not take effect for UEs that report short-TTI capability", "UE capability filter"),
        ("13", "Impact", "TDD", "DL 2-layer MIMO based on TM9 / Massive MIMO", "CellAlgoSwitch.EnhMIMOSwitch → TM9Switch", "If CSI-RS on SCC without SRS, CA UE can use TM9wPMI based on CSI-RS", "See also Ch.19.2"),
        ("14", "Impact", "FDD/TDD", "Energy saving based on proactive scheduling + precise CA scheduling + ultra-low-latency intra-eNB CA", "SymbolPwrSaving.TrigBndlSchDlAvgPrbThld and related ES / precise-scheduling switches", "If PCell not heavily loaded, ultra-low-latency split + ES delay can lower CA UE DL perceived rate", "Ch.5.3.3.3 last RAN row"),
        ("15", "Impact", "FDD/TDD", "VoLTE concurrent with CA", "CaMgtCfg.VolteCaA2RsrpThld; CellQciPara.QciAlgoSwitch → QCI_VOLTE_CA_COEXIST_SW; CaMgtCfg.CellCaAlgoSwitch → VolteSupportCaInterFreqMeasSw", "ACK mode change raises PUSCH bits; MCS lowered to hold IBLER; cell-edge VoLTE uses more RBs. PCC anchoring / SCell config have extra QCI1 conditions", "Related VoLTE parameters grouped here (Ch.19.7)"),
        ("16", "Impact", "FDD/TDD", "PCC anchoring prohibition for high-mobility UEs", "CellOpHoCfg.HighMobiUeHoForbidSw → PCC_ANCHOR_HO_FORBID_SW", "Speed ≤ 30 km/h required for PCC anchoring. NLOS may mis-classify UEs. Lowers E-RAB drop, RRC re-est due to reconf fail, and CA HO counters", "Optional in MML; related to EnhancedPccAnchorSwitch"),
        ("17", "Constraint", "FDD/TDD", "UE / EPC", "UE Rel-10+; supportedBandCombination IE; EPC UE-AMBR/MBR ≥ theoretical peak", "Otherwise peak in Table 5-2 cannot be reached", "Ch.5.3.6"),
        ("18", "Constraint", "TDD", "UL/DL and special subframe", "UL/DL configuration 1 or 2; special subframe 4,5,6,7,9; SS=4 cannot mix with 5/6/7/9; same CP length", "Cells that violate this cannot aggregate", "Ch.5.3.5"),
    ]
    for rec in impacts:
        vals = list(rec) + [""] * 3
        rel = rec[1]
        fh = PALE_GREEN if rel == "Prerequisite" else (PALE_RED if rel == "Exclusive" else (PALE_GOLD if rel == "Constraint" else alt_fill(int(rec[0]))))
        r = table_row(ws, r, vals, fills=[fh] * 10, height=48)
        merge(ws, r - 1, 7, r - 1, COLS)

    # E parameters
    r = section(ws, r, COLS, "E.  All parameters  (SN, MO, name, default, recommended, description, notes)")
    r = note_bar(ws, r, COLS, "Order follows the FPD: group activation → group optimization → adaptive activation → adaptive optimization → common eNodeB switches → common CaMgtCfg. "
                 "Rows with gold fill are related parameters placed immediately under the parent (as requested). "
                 "Default = factory default when stated in the FPD; otherwise “See Parameter Reference”. Recommended = document MML example or explicit “recommended” text.")
    r = subsection(ws, r, COLS, "E1.  CA-group-based — activation  (Table 5-5)")
    r = add_param_header(ws, r)
    r = add_param(ws, r, 1, "ENodeBAlgoSwitch", "CaAlgoSwitch / FreqCfgSwitch", "See Parameter Reference", "Deselect (0)",
                  "Enables group-based mode when FreqCfgSwitch is OFF.",
                  "Parent of all group-based CA. Related options of the SAME parameter are listed immediately below.")
    r = add_param(ws, r, 2, "ENodeBAlgoSwitch", "CaAlgoSwitch / SccBlindCfgSwitch", "See Parameter Reference", "1 if blind SCell is intended, else 0",
                  "Allows blind SCell configuration in group mode.",
                  "Related to FreqCfgSwitch (same MO). Also requires CaGroupSCellCfg.SCellBlindCfgFlag=TRUE (row 12).", related=True)
    r = add_param(ws, r, 3, "ENodeBAlgoSwitch", "CaAlgoSwitch / EnhancedPccAnchorSwitch", "See Parameter Reference", "1 (recommended)",
                  "Enables PCC anchoring for RRC_CONNECTED UEs.",
                  "If both this and PccAnchorSwitch are ON, only Enhanced takes effect. Related: PccSmartCfgSwitch needs this ON.")
    r = add_param(ws, r, 4, "ENodeBAlgoSwitch", "CaAlgoSwitch / SccA2RmvSwitch", "See Parameter Reference", "1",
                  "Deliver A2 measurements for SCells added by A4, so they can be removed.",
                  "Related to CarrAggrA2ThdRsrp / SCellA2Offset (leaving).", related=True)
    r = add_param(ws, r, 5, "ENodeBAlgoSwitch", "CaAlgoSwitch / HoWithSccCfgSwitch", "See Parameter Reference", "0 in doc MML",
                  "Keep SCells during PCC-anchoring HO.",
                  "ON: PCC anchor without removing SCells. OFF: remove SCells first.")
    r = add_param(ws, r, 6, "ENodeBAlgoSwitch", "CaAlgoSwitch / SccModA6Switch", "See Parameter Reference", "0 group / 1 adaptive (doc MML)",
                  "Allow SCell change to a better intra-frequency neighbor (event A6).",
                  "Related: CarrAggrA6Offset, CaA6TimeToTrigger. Group MML example sets 0; adaptive example sets 1.", related=True)
    r = add_param(ws, r, 7, "CaGroup", "CaGroupId", "Site-planned", "0 (doc example)", "Identity of the CA group.", "Max 9 cells per group.")
    r = add_param(ws, r, 8, "CaGroup", "CaGroupTypeInd", "Site-planned", "FDD or TDD (or FDDTDD)", "Group duplex type.", "Doc FDD example: FDD. TDD example: TDD.")
    r = add_param(ws, r, 9, "CaGroupCell", "CaGroupId / eNodeBId / LocalCellId", "Site-planned", "eNodeBId=1234, LocalCellId=0/1 (doc)",
                  "Members of the group.", "Each member is a candidate PCell.")
    r = add_param(ws, r, 10, "CaGroupCell", "PreferredPCellPriority", "See Parameter Reference", "Coverage layer higher (doc: Cell0=2, Cell1=1)",
                  "Higher value → more likely to be PCell.", "Intra-frequency cells should share the same value (eNodeB uses the max).")
    r = add_param(ws, r, 11, "CaGroupCell", "PCellA4RsrpThd / PCellA4RsrqThd", "See Parameter Reference", "−105 dBm / −20 dB (doc MML)",
                  "A5 Thresh2 during group PCC anchoring (named A4 in the MO).",
                  "Related: IntraRatHoComm.InterFreqHoA4TrigQuan chooses RSRP/RSRQ/BOTH. A5 Thresh1 is fixed −43 dBm / −3 dB.", related=True)
    r = add_param(ws, r, 12, "CaGroupSCellCfg", "SCelleNodeBId / SCellLocalCellId", "Site-planned", "1234 / peer LocalCellId (doc)",
                  "Candidate SCell identity.", "Must be in the same CA group as the PCell. Must be inter-freq neighbor except first blind add.")
    r = add_param(ws, r, 13, "CaGroupSCellCfg", "SCellPriority", "See Parameter Reference", "2 or 3 (doc); 0 = never SCell",
                  "Higher → more likely SCell for that PCell.", "Intra-frequency candidates should share the same priority.")
    r = add_param(ws, r, 14, "CaGroupSCellCfg", "SCellBlindCfgFlag", "FALSE typical", "TRUE in doc MML if blind add wanted",
                  "Blind-configure this candidate (no A4).",
                  "Related: SccBlindCfgSwitch must be 1 in group mode. Adaptive: only ONE TRUE per SCC frequency.", related=True)
    r = add_param(ws, r, 15, "CaGroupSCellCfg", "SCellA4Offset", "0 typical", "0 (doc MML)",
                  "Offset added to CarrAggrA4ThdRsrp for this SCell’s A4 add threshold.",
                  "Greater offset → harder to add this SCell. Related parent: CarrAggrA4ThdRsrp.", related=True)
    r = add_param(ws, r, 16, "CaGroupSCellCfg", "SCellA2Offset", "0 typical", "0 (doc MML)",
                  "Offset added to CarrAggrA2ThdRsrp for this SCell’s A2 leave threshold.",
                  "Greater offset → easier to remove this SCell. Related parent: CarrAggrA2ThdRsrp.", related=True)
    r = add_param(ws, r, 17, "CaGroupSCellCfg", "SpidGrpId", "65535 = no restriction", "65535 unless SPID policy",
                  "If ≠ 65535, this candidate cannot be SCell for that SPID group.", "Related SPID policy: SccFreqCfg.SpidGrpId / UlSpidGrpId in adaptive.", related=True)

    r = subsection(ws, r, COLS, "E2.  Adaptive — activation  (Table 5-7)")
    r = add_param_header(ws, r)
    r = add_param(ws, r, 18, "ENodeBAlgoSwitch", "CaAlgoSwitch / FreqCfgSwitch", "See Parameter Reference", "1",
                  "ON selects frequency-based (adaptive) mode.", "Must be paired with AdpCaSwitch.")
    r = add_param(ws, r, 19, "ENodeBAlgoSwitch", "CaAlgoSwitch / AdpCaSwitch", "See Parameter Reference", "1",
                  "Adaptive CA master switch.", "Related: FreqCfgSwitch. Both required.", related=True)
    r = add_param(ws, r, 20, "CaMgtCfg", "CellCaAlgoSwitch / CaDl2CCExtSwitch", "OFF", "1 when aggregated BW is in the 40 MHz window (FDD >20, TDD >30)",
                  "Extends 2CC to 40 MHz aggregated bandwidth in adaptive mode.",
                  "Consumes LAOFD-001002 / TDLAOFD-001002. Group mode does not need this bit.")
    r = add_param(ws, r, 21, "ENodeBAlgoSwitch", "CaAlgoExtSwitch / FreqBaseCaLicAlarmSwitch", "See Parameter Reference", "1 (doc)",
                  "Frequency-based CA license alarm.", "Turn on in adaptive MML examples.")
    r = add_param(ws, r, 22, "PccFreqCfg", "PccDlEarfcn", "Site-planned", "FDD example 123/456; TDD example 37900/38098",
                  "Candidate PCC downlink EARFCN.", "If no SccFreqCfg for this PCC, A5 for PCC anchoring is not sent.")
    r = add_param(ws, r, 23, "PccFreqCfg", "PreferredPccPriority", "See Parameter Reference", "1 and 2 (doc)",
                  "Higher → more likely PCC.", "Related to EnhancedPccAnchorSwitch.")
    r = add_param(ws, r, 24, "PccFreqCfg", "PccA4RsrpThd / PccA4RsrqThd", "See Parameter Reference", "−105 dBm / −20 dB (doc)",
                  "A5 Thresh2 for adaptive PCC anchoring.",
                  "Recommended: PccA4RsrpThd > InterFreqHoGroup.InterFreqHoA4ThdRsrp so coverage HO success rate does not drop.", related=True)
    r = add_param(ws, r, 25, "SccFreqCfg", "PccDlEarfcn + SccDlEarfcn", "Site-planned", "Pair each PCC with the other EARFCN (doc)",
                  "Candidate SCC under a PCC.", "Total PCC+SCC frequencies ≤ 17.")
    r = add_param(ws, r, 26, "SccFreqCfg", "SccPriority", "See Parameter Reference", "2 or 3 (doc)",
                  "Higher → more likely SCC for that PCC.", "Flexible CA OFF uses this order to fill A4 objects.")
    r = add_param(ws, r, 27, "SccFreqCfg", "SccA4Offset / SccA2Offset", "0 typical", "0 (doc)",
                  "Per-SCC add/leave offsets vs CarrAggrA4/A2ThdRsrp.",
                  "A2 sum is clamped to [−140, −43] dBm. Related parents: CarrAggrA4ThdRsrp, CarrAggrA2ThdRsrp.", related=True)
    r = add_param(ws, r, 28, "SccFreqCfg", "CnOperatorList", "Site-planned", '"111111" (doc)',
                  "Operators allowed to use this SCC.", "MOCN: set to serving operators.")
    r = add_param(ws, r, 29, "CaMgtCfg", "SCellAgingTime", "See Parameter Reference", "15 (doc adaptive MML)",
                  "Seconds of no CA between a pair before the dynamic route is released.",
                  "If route capacity is tight, use a smaller value. Related: CaRouteConfigPenaltyOfs / CaRouteConfigPenaltyWeight.", related=True)

    r = subsection(ws, r, COLS, "E3.  Common algorithm switches  (Ch.5.4.1.1.3)  —  both modes")
    r = add_param_header(ws, r)
    r = add_param(ws, r, 30, "ENodeBAlgoSwitch", "CaAlgoSwitch / PccAnchorSwitch", "See Parameter Reference", "Deselect (Enhanced is preferred)",
                  "Legacy PCC anchoring.", "If Enhanced is ON, this is ignored.")
    r = add_param(ws, r, 31, "ENodeBAlgoSwitch", "CaAlgoSwitch / CaTrafficTriggerSwitch", "See Parameter Reference", "Select only if traffic-based add/remove is wanted",
                  "SCell config/remove based on traffic; also allows PCC anchoring on unnecessary HO.",
                  "ON: blind SCell does not take effect; after traffic deactivation SCells are RRC-removed. Ping-pong PCC risk.")
    r = add_param(ws, r, 32, "ENodeBAlgoSwitch", "CaAlgoSwitch / PccSmartCfgSwitch", "See Parameter Reference", "Select for load-based PCell",
                  "Skip high-load candidate PCells.",
                  "Requires PccAnchorSwitch or EnhancedPccAnchorSwitch. High load = PCell UEs > CellMaxPccNumber × PccUserNumberOffloadThd, or BBP CPU > 70%.", related=True)
    r = add_param(ws, r, 33, "ENodeBAlgoSwitch", "CaAlgoSwitch / SccSmartCfgSwitch", "See Parameter Reference", "Select for load-based SCell",
                  "Skip high-load candidate SCells.", "Companion of PccSmartCfgSwitch. Related: HLUeCntThldForScellConfig.", related=True)
    r = add_param(ws, r, 34, "ENodeBAlgoSwitch", "CaAlgoSwitch / IdleModePccAnchorSwitch", "See Parameter Reference", "Select if IDLE PCC anchoring wanted",
                  "PCC anchoring for RRC_IDLE UEs (Ch.4.7).", "")
    r = add_param(ws, r, 35, "ENodeBAlgoSwitch", "CaAlgoSwitch / CaAdpPreSchSwitch", "Selected (doc default)", "Keep selected in listed scenarios",
                  "Adaptive pre-scheduling to raise single-CA-UE throughput (relaxed BH, distributed/hybrid coordination FDD, TDD 2CC enhance / ≥3CC / Dl2CCAckResShareSw).",
                  "Document: default status is selected.")
    r = add_param(ws, r, 36, "ENodeBAlgoSwitch", "CaAlgoExtSwitch / CaA5HoEventSwitch (+ CaA5HoEventEnhSwitch)", "See Parameter Reference", "Select BOTH (doc recommendation)",
                  "Change HO event A4 to A5 so PCell can HO to an SCell.",
                  "Related: InterFreqHoGroup.InterFreqHoA5Thd1Rsrp/Rsrq. Without this, SCell cannot become PCell via A4.", related=True)
    r = add_param(ws, r, 37, "ENodeBAlgoSwitch", "CaAlgoExtSwitch / CaMubfPairingAdaptOptSwitch", "See Parameter Reference", "Optional (doc TDD/FDD adaptive MML)",
                  "Adaptive handling of CA for MU beamforming.",
                  "Requires CaSmartSelectionSwitch OFF. Related: HLUeCntThldForScellConfig, HighLoadCellTypeNotAsScell, RSVD_SW_PARAM0_BIT28 (TDD MM).", related=True)
    r = add_param(ws, r, 38, "ENodeBAlgoSwitch", "CaAlgoExtSwitch / CaEnhancedPreAllocSwitch", "See Parameter Reference", "Select on moderate/light MBB if Dl2CCAckResShareSw is ON",
                  "Dynamically split traffic across CCs by real-time volume and scheduling capability.",
                  "Cannot be ON together with DlCaLbAlgoSwitch or CaLoadBalancePreAllocSwitch. FDD relaxed-BH: no effect. Raises L.Thrp.bits.DL.CAUser / L.Thrp.Time.DL.CAUser.")
    r = add_param(ws, r, 39, "ENodeBAlgoSwitch", "CaLbAlgoSwitch / DlCaLbAlgoSwitch", "See Parameter Reference", "Select for inter-CC load transfer",
                  "Inter-CC load transfer triggered by cell load.", "Mutually exclusive with CaEnhancedPreAllocSwitch.", related=True)
    r = add_param(ws, r, 40, "EnodebCounterParaGrp", "EnodebCounterAlgoSwitch / CA_UE_MULTI_CC_STAT_OPT_SW", "OFF", "1 (doc MML — turn ON)",
                  "Count a UE as nCC-active as soon as n carriers are active (not requiring ALL SCells active).",
                  "Changes KPI stepping of L.Traffic.User.PCell.DL.Active.* and L.Thrp.*.nCC.CAUser.")
    r = add_param(ws, r, 41, "GlobalProcSwitch", "ProtocolCompatibilitySw / SCellBlindA2Switch", "See Parameter Reference", "1 in group MML",
                  "Protocol compatibility for blind SCell A2.", "Group activation example.")
    r = add_param(ws, r, 42, "GlobalProcSwitch", "ProtocolCompatibilitySw / SCellModCaMeasRmvSwitch", "See Parameter Reference", "1 (doc)",
                  "Compatibility for CA measurement removal on SCell change.", "Both group and adaptive examples.")
    r = add_param(ws, r, 43, "CellSiMap", "SiSwitch / ForbidCellSiSwitch", "OFF", "Select only to make a dedicated SCell",
                  "Cell cannot be PCell; UEs cannot camp/access/HO to it.",
                  "Related: CaMgtCfg.CellCaAlgoSwitch / UlScellForbidSwitch (band 38 DL-only SCell) takes effect only if this is ON.", related=True)
    r = add_param(ws, r, 44, "CellAlgoSwitch", "PucchAlgoSwitch / Dl2CCAckResShareSw", "See Parameter Reference", "Select to expand CA route capacity (with CaRouteNumberExtensionSwitch)",
                  "ACK resource sharing; more UEs enter CA but per-UE rate may drop.",
                  "Required before CaRouteNumberExtensionSwitch. 8 / 24 / 48 neighbor routes. 48 needs UMPT+UBBP, CPU rises.", related=True)
    r = add_param(ws, r, 45, "CaMgtCfg", "CellCaAlgoSwitch / CaRouteNumberExtensionSwitch", "OFF", "Select with Dl2CCAckResShareSw if >24 routes needed",
                  "Extend per-cell CA route number.", "Cannot be ON if Dl2CCAckResShareSw is OFF.", related=True)
    r = add_param(ws, r, 46, "CellPucchAlgo", "SriAlgoSwitch / SriPeriodCfgOptSW", "See Parameter Reference", "Select (doc recommendation)",
                  "SRI period optimisation.", "If OFF and SRI period is long, RAB setup success and drop rate may worsen.")

    r = subsection(ws, r, COLS, "E4.  Common CaMgtCfg carrier-management parameters  (add / leave / activate / deactivate)")
    r = add_param_header(ws, r)
    r = add_param(ws, r, 47, "CaMgtCfg", "CarrAggrA4ThdRsrp", "See Parameter Reference", "Set > CarrAggrA2ThdRsrp; typical same order as −105 (doc A4 examples)",
                  "Base RSRP threshold for CA event A4 (SCell add).",
                  "MUST be greater than CarrAggrA2ThdRsrp. Children: SCellA4Offset, SccA4Offset.", related=True)
    r = add_param(ws, r, 48, "CaMgtCfg", "CaA4TimeToTrigger", "See Parameter Reference", "Site (balance speed vs ping-pong)",
                  "Time-to-trigger for CA A4.", "Larger → fewer A4 reports.")
    r = add_param(ws, r, 49, "CaMgtCfg", "CarrAggrA2ThdRsrp", "See Parameter Reference", "Lower than A4 (leave hysteresis), e.g. several dB below A4",
                  "Base RSRP threshold for CA event A2 (SCell remove).",
                  "Children: SCellA2Offset, SccA2Offset. Greater value → more removals.", related=True)
    r = add_param(ws, r, 50, "CaMgtCfg", "CaA2TimeToTrigger", "See Parameter Reference", "Site",
                  "Time-to-trigger for CA A2.", "Larger → fewer removals.")
    r = add_param(ws, r, 51, "CaMgtCfg", "CarrAggrA6Offset", "See Parameter Reference", "Small positive dB (example 3 in worked sheet)",
                  "Offset Off in event A6 (SCell change).",
                  "Greater → fewer SCell changes. Related: SccModA6Switch, CaA6TimeToTrigger (must be ≤ 3 s).", related=True)
    r = add_param(ws, r, 52, "CaMgtCfg", "CaA6TimeToTrigger / CaA6ReportAmount / CaA6ReportInterval", "See Parameter Reference", "TTT ≤ 3 s",
                  "A6 TTT, number of reports, interval.", "If TTT > 3 s, A6 cannot be reported (gap timer).")
    r = add_param(ws, r, 53, "CaMgtCfg", "CarrierMgtSwitch", "See Parameter Reference", "ON to allow traffic/CQI deactivation",
                  "Master switch for SCell deactivation by traffic or channel quality.",
                  "OFF: SCells deactivate only on RLF. Related: Deactive* and SccDeactCqiThd.", related=True)
    r = add_param(ws, r, 54, "CaMgtCfg", "ActiveBufferDelayThd", "See Parameter Reference", "Site; smaller → faster activation",
                  "Buffer delay threshold for SCell activation.", "NSA uses NsaDcLteScellActBfrDelThld instead.")
    r = add_param(ws, r, 55, "CaMgtCfg", "ActiveBufferLenThd", "See Parameter Reference", "Non-zero; MUST be > DeactiveBufferLenThd",
                  "Buffer length threshold for SCell activation.",
                  "Do NOT set 0 (immediate activate + immediate deactivate → BBP overload). Related: DeactiveBufferLenThd.", related=True)
    r = add_param(ws, r, 56, "CaMgtCfg", "DeactiveThroughputThd", "See Parameter Reference", "Site; greater → easier deactivate",
                  "Uu rate threshold for traffic deactivation.", "Related: CarrierMgtSwitch.", related=True)
    r = add_param(ws, r, 57, "CaMgtCfg", "DeactiveBufferLenThd", "See Parameter Reference", "Less than ActiveBufferLenThd (doc requirement)",
                  "Buffer length threshold for traffic deactivation.", "Hysteresis vs row 55.")
    r = add_param(ws, r, 58, "CaMgtCfg", "SccDeactCqiThd", "0 = disabled", "5 (explicitly recommended)",
                  "CQI threshold for channel-quality deactivation (non-relaxed BH).",
                  "0 disables. Relaxed BH uses RelaxedBHSccDeactCqiThd. Related: SccReactivationTime, CaSccSuspendSwitch must be OFF.", related=True)
    r = add_param(ws, r, 59, "CaMgtCfg", "SccCfgInterval", "See Parameter Reference", "Smaller to retry SCell add faster",
                  "Min interval to retry SCell configuration after failure, while traffic still meets activation.",
                  "Smaller → more RRC Reconfigurations and more gap-measurement throughput loss.")
    r = add_param(ws, r, 60, "CaMgtCfg", "CellMaxPccNumber", "See Parameter Reference", "Match BBP/PCell capacity",
                  "Max PCell-served CA UEs in the cell.",
                  "Hysteresis 90% to resume. Related: PccUserNumberOffloadThd for load-based PCC.", related=True)
    r = add_param(ws, r, 61, "CaMgtCfg", "Forbid2ccUeRatioThld", "100 (function OFF)", "90 if UE-number control needed (doc optional MML)",
                  "UE-proportion threshold to forbid 2CC or higher.",
                  "Must be > Forbid3/4/5 when those ≠ 100. Related children below.", related=True)
    r = add_param(ws, r, 62, "CaMgtCfg", "Forbid3ccUeRatioThld", "100", "80 (doc optional MML)", "Forbid 3CC or higher.", "Must be < Forbid2 and > Forbid4/5.", related=True)
    r = add_param(ws, r, 63, "CaMgtCfg", "Forbid4ccUeRatioThld", "100", "70 (doc optional MML)", "Forbid 4CC or higher.", "Used also in Step3.", related=True)
    r = add_param(ws, r, 64, "CaMgtCfg", "Forbid5ccUeRatioThld", "100", "60 (doc optional MML)", "Forbid 5CC or higher.", "Used also in Step4.", related=True)
    r = add_param(ws, r, 65, "CaMgtCfg", "CellCaAlgoSwitch / CaInstantlyJudgeSwitch", "OFF", "ON for small-packet (web) faster SCell activate; OFF for large-file",
                  "Instant ms vs filtered s traffic for activation.", "ON can raise CA UE ratio and IBLER/RBLER fluctuation.")
    r = add_param(ws, r, 66, "CaMgtCfg", "CellCaAlgoSwitch / 2CCDlCaEnhanceSwitch", "OFF", "ON if PUCCH format 3 wanted for 2CC UEs that support it",
                  "Allocate PUCCH format 3 for 2CC UEs (2 RBs if BW≥10 MHz, 1 RB if BW≤5 MHz).",
                  "UE with ≥3CC capability is treated as PUCCH format-3 capable. Causes PCell IBLER fluctuation.")
    r = add_param(ws, r, 67, "CaMgtCfg", "ScellNoActivationUeNumThld", "0 = OFF", "Non-zero on FDD 1.4/3/5 MHz and TDD 5 MHz",
                  "Heavy-load SCell activation prohibit.", "Not compatible with LBBPc.")
    r = add_param(ws, r, 68, "CaMgtCfg", "EnhancedPccAnchorA1ThdRsrp", "See Parameter Reference", "Site; example −90 dBm in A1 worked example",
                  "A1 RSRP threshold that starts PCC anchoring.", "Related: EnhancedPccAnchorSwitch.")
    r = add_param(ws, r, 69, "CaMgtCfg", "HLUeCntThldForScellConfig / HighLoadCellTypeNotAsScell", "See Parameter Reference", "100 / MASSIVE_MIMO_CELL or NORMAL_CELL (doc optional)",
                  "Load-based SCell exclusion.", "Used with CaMubfPairingAdaptOptSwitch.")
    r = add_param(ws, r, 70, "CellOpHoCfg", "HighMobiUeHoForbidSw / PCC_ANCHOR_HO_FORBID_SW", "OFF", "1 if high-mobility UEs must not PCC-anchor (doc optional)",
                  "PCC anchoring only if UE speed ≤ 30 km/h.", "NLOS false high-mobility risk.")
    r = add_param(ws, r, 71, "EutranInterNFreq / EutranInterFreqNCell", "DlEarfcn, MeasBandWidth, NoRmvFlag", "Site", "NoRmvFlag=FORBID_RMV_ENUM (mandatory in doc)",
                  "Inter-frequency neighbors for CA.",
                  "If ANR can remove the neighbor, CA stops. Related: ADD EUTRANINTERNFREQ / INTERFREQNCELL MML.", related=True)
    r = add_param(ws, r, 72, "CqiAdaptiveCfg", "SimulAckNackAndCqiFmt3Sw", "See Parameter Reference", "ON if simultaneous ACK/NACK + periodic CQI on format 3 wanted (3CC+)",
                  "Used by 3CC/4CC/5CC. Listed here because 2CC enhance also uses PUCCH format 3.", "If OFF, ACK/NACK and periodic CQI cannot share the subframe.")

    # F MML
    r = section(ws, r, COLS, "F.  MML commands  (all examples from Ch.5.4.1.2, in document sequence, with values)")
    r = note_bar(ws, r, COLS, "Values are exactly those printed in the FPD (eNodeBId=1234, MCC/MNC 460/20, EARFCN 37900/38098 or 123/456). "
                 "Replace with live-network IDs. Always complete Ch.5.3 requirements first. Check “Service Interrupted After Modification” in the Parameter Reference.")
    r = add_mml_header(ws, r)
    mmls = [
        (1, "0.1", "Common", "FDD/TDD", "Pre-step — neighbor freq",
         'ADD EUTRANINTERNFREQ: LocalCellId=0, DlEarfcn=37900, UlEarfcnCfgInd=NOT_CFG, CellReselPriorityCfgInd=NOT_CFG, SpeedDependSPCfgInd=NOT_CFG, MeasBandWidth=MBW100, PmaxCfgInd=NOT_CFG, QqualMinCfgInd=NOT_CFG;',
         "Repeat for LocalCellId=1, DlEarfcn=38098."),
        (2, "0.2", "Common", "FDD/TDD", "Pre-step — neighbor freq",
         'ADD EUTRANINTERNFREQ: LocalCellId=1, DlEarfcn=38098, UlEarfcnCfgInd=NOT_CFG, CellReselPriorityCfgInd=NOT_CFG, SpeedDependSPCfgInd=NOT_CFG, MeasBandWidth=MBW100, PmaxCfgInd=NOT_CFG, QqualMinCfgInd=NOT_CFG;',
         "Pair of frequencies."),
        (3, "0.3", "Common", "FDD/TDD", "Pre-step — external cell (inter-eNB only)",
         'ADD EUTRANEXTERNALCELL: Mcc="460", Mnc="20", eNodeBId=1234, CellId=1, DlEarfcn=37900, UlEarfcnCfgInd=NOT_CFG, PhyCellId=101, Tac=1;',
         "Optional. Also CellId=0, DlEarfcn=38098."),
        (4, "0.4", "Common", "FDD/TDD", "Pre-step — RAN sharing PLMN (optional)",
         'ADD EUTRANEXTERNALCELLPLMN: Mcc="460", Mnc="20", eNodeBId=1234, CellId=1, ShareMcc="460", ShareMnc="22";',
         "Optional. Repeat CellId=0."),
        (5, "0.5", "Common", "FDD/TDD", "Pre-step — inter-freq neighbor (MANDATORY NoRmvFlag)",
         'ADD EUTRANINTERFREQNCELL: LocalCellId=0, Mcc="460", Mnc="20", eNodeBId=1234, CellId=1, NoRmvFlag=FORBID_RMV_ENUM;',
         "Without FORBID_RMV_ENUM, ANR may delete the relation and CA stops. Reverse: LocalCellId=1, CellId=0."),
        (6, "0.6", "Common", "FDD/TDD", "Pre-step — intra-freq neighbor",
         'ADD EUTRANINTRAFREQNCELL: LocalCellId=0, Mcc="460", Mnc="20", eNodeBId=1234, CellId=2;',
         "And reverse LocalCellId=2, CellId=0. Needed for A6 SCell change."),
        (7, "1.1", "Group", "FDD", "Activation — switches",
         'MOD ENODEBALGOSWITCH: CaAlgoSwitch=SccBlindCfgSwitch-1&FreqCfgSwitch-0&SccA2RmvSwitch-1&HoWithSccCfgSwitch-0&SccModA6Switch-0&EnhancedPccAnchorSwitch-1;',
         "Ch.5.4.1.2.1. FreqCfgSwitch-0 = group mode."),
        (8, "1.2", "Group", "FDD", "Activation — protocol",
         'MOD GLOBALPROCSWITCH: ProtocolCompatibilitySw=SCellBlindA2Switch-1&SCellModCaMeasRmvSwitch-1;',
         ""),
        (9, "1.3", "Group", "FDD", "Activation — group",
         'ADD CAGROUP: CaGroupId=0, CaGroupTypeInd=FDD;',
         "TDD uses CaGroupTypeInd=TDD (command 1.3b)."),
        (10, "1.4", "Group", "FDD/TDD", "Activation — members",
         'ADD CAGROUPCELL: CaGroupId=0, LocalCellId=0, eNodeBId=1234, PreferredPCellPriority=2, PCellA4RsrpThd=-105, PCellA4RsrqThd=-20;',
         "Cell1: PreferredPCellPriority=1, same A4 thresholds."),
        (11, "1.5", "Group", "FDD/TDD", "Activation — members",
         'ADD CAGROUPCELL: CaGroupId=0, LocalCellId=1, eNodeBId=1234, PreferredPCellPriority=1, PCellA4RsrpThd=-105, PCellA4RsrqThd=-20;',
         ""),
        (12, "1.6", "Group", "FDD/TDD", "Activation — candidate SCells",
         'ADD CAGROUPSCELLCFG: LocalCellId=0, SCelleNodeBId=1234, SCellLocalCellId=1, SCellBlindCfgFlag=TRUE, SCellPriority=2, SCellA4Offset=0, SCellA2Offset=0;',
         "Must already be inter-freq neighbor. Reverse: LocalCellId=1, SCellLocalCellId=0, SCellPriority=3."),
        (13, "1.7", "Group", "FDD/TDD", "Activation — candidate SCells",
         'ADD CAGROUPSCELLCFG: LocalCellId=1, SCelleNodeBId=1234, SCellLocalCellId=0, SCellBlindCfgFlag=TRUE, SCellPriority=3, SCellA4Offset=0, SCellA2Offset=0;',
         ""),
        (14, "1.8", "Group", "FDD/TDD", "Activation — counter opt",
         'MOD ENODEBCOUNTERPARAGRP: EnodebCounterAlgoSwitch=CA_UE_MULTI_CC_STAT_OPT_SW-1;',
         "Affects KPI formulas below."),
        (15, "1.9", "Group", "FDD/TDD", "Optional — high-mobility PCC forbid",
         'MOD CELLOPHOCFG: LocalCellId=0, CnOperatorId=0, HighMobiUeHoForbidSw=PCC_ANCHOR_HO_FORBID_SW-1;',
         "Repeat LocalCellId=1."),
        (16, "1.10", "Group", "FDD/TDD", "Optional — UE-number SCell control",
         'MOD CAMGTCFG: LocalCellId=0, Forbid5ccUeRatioThld=60, Forbid4ccUeRatioThld=70, Forbid3ccUeRatioThld=80, Forbid2ccUeRatioThld=90;',
         "Repeat LocalCellId=1. Restore to 100 to disable."),
        (17, "1.11", "Group", "FDD/TDD", "Activation — cell up",
         'ACT CELL: LocalCellId=0;',
         "Also ACT CELL: LocalCellId=1;"),
        (18, "1.12", "Group", "FDD/TDD", "Deactivation — switches",
         'MOD ENODEBALGOSWITCH: CaAlgoSwitch=SccBlindCfgSwitch-0&FreqCfgSwitch-0&SccA2RmvSwitch-0&HoWithSccCfgSwitch-0&SccModA6Switch-0&EnhancedPccAnchorSwitch-0;',
         "Then GLOBALPROCSWITCH bits 0, counter SW 0, CELLOPHOCFG 0, Forbid* =100."),
        (19, "1.13", "Group", "FDD/TDD", "Deactivation — remove group",
         'RMV CAGROUPCELL: CaGroupId=0, LocalCellId=0, eNodeBId=1234;',
         "Also LocalCellId=1, then RMV CAGROUP: CaGroupId=0;"),
        (20, "2.1", "Adaptive", "FDD", "Activation — switches",
         'MOD ENODEBALGOSWITCH: CaAlgoSwitch=SccBlindCfgSwitch-0&FreqCfgSwitch-1&SccA2RmvSwitch-1&HoWithSccCfgSwitch-0&SccModA6Switch-1&AdpCaSwitch-1&EnhancedPccAnchorSwitch-1;',
         "Ch.5.4.1.2.3. Note SccModA6Switch-1 vs group 0."),
        (21, "2.2", "Adaptive", "FDD", "Activation — protocol",
         'MOD GLOBALPROCSWITCH: ProtocolCompatibilitySw=SCellModCaMeasRmvSwitch-1;',
         "No SCellBlindA2Switch in adaptive example."),
        (22, "2.3", "Adaptive", "FDD", "Activation — PCC",
         'ADD PCCFREQCFG: PccDlEarfcn=123, PreferredPccPriority=1, PccA4RsrpThd=-105, PccA4RsrqThd=-20;',
         "Second: PccDlEarfcn=456, PreferredPccPriority=2, same A4."),
        (23, "2.4", "Adaptive", "FDD", "Activation — PCC",
         'ADD PCCFREQCFG: PccDlEarfcn=456, PreferredPccPriority=2, PccA4RsrpThd=-105, PccA4RsrqThd=-20;',
         ""),
        (24, "2.5", "Adaptive", "FDD", "Activation — SCC",
         'ADD SCCFREQCFG: PccDlEarfcn=123, SccDlEarfcn=567, SccPriority=2, SccA2Offset=0, SccA4Offset=0, CnOperatorList="111111";',
         "Second pair: 456 / 789, SccPriority=3."),
        (25, "2.6", "Adaptive", "FDD", "Activation — SCC",
         'ADD SCCFREQCFG: PccDlEarfcn=456, SccDlEarfcn=789, SccPriority=3, SccA2Offset=0, SccA4Offset=0, CnOperatorList="111111";',
         ""),
        (26, "2.7", "Adaptive", "FDD/TDD", "Optional — blind flag",
         'ADD CAGROUPSCELLCFG: LocalCellId=0, SCelleNodeBId=1234, SCellLocalCellId=1, SCellBlindCfgFlag=TRUE;',
         "Only one TRUE per SCC frequency. Reverse LocalCellId=1 / SCellLocalCellId=0."),
        (27, "2.8", "Adaptive", "FDD/TDD", "Activation — license alarm",
         'MOD ENODEBALGOSWITCH: CaAlgoExtSwitch=FreqBaseCaLicAlarmSwitch-1;',
         ""),
        (28, "2.9", "Adaptive", "FDD", "Optional — 40 MHz window",
         'MOD CAMGTCFG: LocalCellId=0, SCellAgingTime=15, CellCaAlgoSwitch=CaDl2CCExtSwitch-1;',
         "FDD: when 20 < BW ≤ 40 MHz. Repeat LocalCellId=1. TDD: when 30 < BW ≤ 40 MHz (same command)."),
        (29, "2.10", "Adaptive", "TDD", "Activation — PCC (TDD EARFCN)",
         'ADD PCCFREQCFG: PccDlEarfcn=37900, PreferredPccPriority=1, PccA4RsrpThd=-105, PccA4RsrqThd=-20;',
         "Ch.5.4.1.2.4. Second: 38098, priority 2."),
        (30, "2.11", "Adaptive", "TDD", "Activation — SCC (TDD EARFCN)",
         'ADD SCCFREQCFG: PccDlEarfcn=37900, SccDlEarfcn=38098, SccPriority=2, SccA2Offset=0, SccA4Offset=0, CnOperatorList="111111";',
         "Reverse: 38098 / 37900, SccPriority=3."),
        (31, "2.12", "Adaptive", "FDD/TDD", "Optional — MU-BF / MM SCell",
         'MOD ENODEBALGOSWITCH: CaAlgoSwitch=FreqCfgSwitch-1&AdpCaSwitch-1&CaSmartSelectionSwitch-0, CaAlgoExtSwitch=CaMubfPairingAdaptOptSwitch-1;',
         "Then MOD CAMGTCFG: HLUeCntThldForScellConfig=100, HighLoadCellTypeNotAsScell=MASSIVE_MIMO_CELL or NORMAL_CELL."),
        (32, "2.13", "Adaptive", "TDD", "Optional — max one high-load MM SCell",
         'MOD ENODEBRSVDPARAMEXT: RsvdSwParam0=RSVD_SW_PARAM0_BIT28-1;',
         "TDD only, with MASSIVE_MIMO_CELL."),
        (33, "2.14", "Adaptive", "FDD/TDD", "Deactivation — remove freq",
         'RMV SCCFREQCFG: PccDlEarfcn=123, SccDlEarfcn=567;',
         "Then RMV PCCFREQCFG. TDD uses 37900/38098."),
        (34, "3.1", "Common", "FDD/TDD", "Verification",
         'DSP UEONLINEINFO;',
         "Check SCell configuration status and last SCell config failure cause (Ch.5.4.2)."),
    ]
    for rec in mmls:
        r = add_mml(ws, r, *rec)

    r = note_bar(ws, r, COLS, "MAE-Deployment alternative: Feature Operation and Maintenance (batch) or Feature Configuration Using the MAE-Deployment (single/batch). GUI may differ by OSS version.")

    # G counters
    r = section(ws, r, COLS, "G.  Counter list to monitor  (Ch.5.4.2 Table 5-9 + Ch.5.4.3 Tables 5-10, 5-11, 5-12)")
    r = add_ctr_header(ws, r)
    ctrs = [
        (1, "1526728426", "L.Traffic.User.PCell.DL.Avg", "Average number of UEs using this cell as DL PCell", "Activation verify (any non-zero among Table 5-9)"),
        (2, "1526728516", "L.Traffic.User.PCell.DL.Max", "Max UEs using this cell as DL PCell", "Activation + monitoring"),
        (3, "1526728427", "L.Traffic.User.SCell.DL.Avg", "Average UEs using this cell as DL SCell", "Activation verify"),
        (4, "1526728517", "L.Traffic.User.SCell.DL.Max", "Max UEs using this cell as DL SCell", "Monitoring"),
        (5, "1526728424", "L.ChMeas.PRB.DL.PCell.Used.Avg", "Average used DL PRBs by PCell CA traffic", "Resource usage"),
        (6, "1526728425", "L.ChMeas.PRB.DL.SCell.Used.Avg", "Average used DL PRBs by SCell CA traffic", "Resource usage"),
        (7, "1526737791", "L.Traffic.User.PCell.DL.Active.Avg", "Avg UEs with SCell activated (see MULTI_CC_STAT_OPT_SW)", "Activation verify"),
        (8, "1526737792", "L.Traffic.User.PCell.DL.Active.Max", "Max UEs with SCell activated", "Activation verify"),
        (9, "1526726740", "L.ChMeas.PRB.DL.Used.Avg", "Overall DL PRB usage", "Compare before/after CA"),
        (10, "1526728564", "L.Thrp.bits.DL.CAUser", "DL PDCP bits of CA UEs", "KPI throughput numerator"),
        (11, "1526728565", "L.Thrp.Time.DL.CAUser", "DL PDCP duration of CA UEs", "KPI throughput denominator"),
        (12, "1526728514", "L.E-RAB.AbnormRel.CAUser", "Abnormal E-RAB releases of CA UEs", "Drop-rate numerator"),
        (13, "1526728515", "L.E-RAB.NormRel.CAUser", "Normal E-RAB releases of CA UEs", "Drop-rate denominator part"),
        (14, "1526728518", "L.HHO.PrepAttOut.CAUser.PCC", "CA UE outgoing HO preparations (PCC)", "HO funnel"),
        (15, "1526728519", "L.HHO.ExecAttOut.CAUser.PCC", "CA UE outgoing HO executions (PCC)", "HO success denominator"),
        (16, "1526728520", "L.HHO.ExecSuccOut.CAUser.PCC", "CA UE outgoing HO success (PCC)", "HO success numerator"),
        (17, "1526729602", "L.HHO.InterFddTdd.PrepAttOut.CAUser.PCC", "FDD↔TDD HO prep of CA PCell", "Multi-RAT HO"),
        (18, "1526729603", "L.HHO.InterFddTdd.ExecAttOut.CAUser.PCC", "FDD↔TDD HO exec of CA PCell", "Multi-RAT HO"),
        (19, "1526729604", "L.HHO.InterFddTdd.ExecSuccOut.CAUser.PCC", "FDD↔TDD HO success of CA PCell", "Multi-RAT HO success"),
        (20, "1526729045", "L.CA.DLSCell.Add.Att", "DL SCell add attempts", "Add success rate"),
        (21, "1526729046", "L.CA.DLSCell.Add.Succ", "DL SCell add success", "Add success rate"),
        (22, "1526729047", "L.CA.DLSCell.Rmv.Att", "DL SCell remove attempts", "Leave procedure"),
        (23, "1526729048", "L.CA.DLSCell.Rmv.Succ", "DL SCell remove success", "Leave procedure"),
        (24, "1526730592", "L.CA.DLSCell.Add.Meas.Att", "Measurement-based (A4) SCell add attempts", "A4 path"),
        (25, "1526730593", "L.CA.DLSCell.Add.Meas.Succ", "Measurement-based SCell add success", "A4 path"),
        (26, "1526730594", "L.CA.DLSCell.Rmv.Meas.Att", "Measurement-based SCell remove attempts", "A2 path"),
        (27, "1526730595", "L.CA.DLSCell.Rmv.Meas.Succ", "Measurement-based SCell remove success", "A2 path"),
        (28, "1526730590", "L.CA.DLSCell.Add.Blind.Att", "Blind SCell add attempts", "Blind path"),
        (29, "1526730591", "L.CA.DLSCell.Add.Blind.Succ", "Blind SCell add success", "Blind path"),
        (30, "1526730596", "L.CA.DLSCell.Mod.Att", "SCell change (A6) attempts", "SCell change"),
        (31, "1526730597", "L.CA.DLSCell.Mod.Succ", "SCell change success", "SCell change"),
        (32, "1526728999", "L.CA.DLSCell.Act.Att", "SCell activation attempts (MAC CE)", "Activation"),
        (33, "1526729000", "L.CA.DLSCell.Act.Succ", "SCell activation success", "Activation"),
        (34, "1526729001", "L.CA.DLSCell.Deact.Att", "SCell deactivation attempts", "Leaving (MAC)"),
        (35, "1526729002", "L.CA.DLSCell.Deact.Succ", "SCell deactivation success", "Leaving (MAC)"),
        (36, "1526732658", "L.CA.Traffic.bits.DL.PCell", "DL bits on PCell of CA UEs", "Split PCell vs SCell traffic"),
        (37, "1526729259", "L.CA.Traffic.bits.DL.SCell", "DL bits on SCell of CA UEs", "Split PCell vs SCell traffic"),
        (38, "1526729003", "L.CA.DL.PCell.Act.Dur", "Active duration on PCell", "Duration KPI"),
        (39, "1526729004", "L.CA.DL.SCell.Act.Dur", "Active duration on SCell", "Duration KPI"),
        (40, "1526732656", "L.Traffic.User.SCell.Active.DL.Avg", "Avg UEs with this cell as activated SCell", "Active SCell users"),
        (41, "1526732657", "L.Traffic.User.SCell.Active.DL.Max", "Max UEs with this cell as activated SCell", "Active SCell users"),
        (42, "1526767257", "L.CA.MeasRpts.InterEnb.Strongest.NCELL.A6.Num", "Inter-eNB strongest-neighbor A6 reports", "Inter-eNB A6"),
        (43, "1526739796", "L.RB.DL.PCell.CAUsed.PLMN", "MOCN: DL RBs used as PCell per PLMN", "MOCN Table 5-12"),
        (44, "1526739797", "L.RB.DL.SCell.CAUsed.PLMN", "MOCN: DL RBs used as SCell per PLMN", "MOCN"),
        (45, "1526739766", "L.Traffic.User.PCell.DL.Avg.PLMN", "MOCN: PCell CA users per PLMN", "MOCN"),
        (46, "1526739767", "L.Traffic.User.SCell.DL.Avg.PLMN", "MOCN: SCell CA users per PLMN", "MOCN"),
        (47, "1526739743", "L.Thrp.bits.DL.CAUser.PLMN", "MOCN: CA UE DL bits per PLMN", "MOCN throughput"),
        (48, "1526739744", "L.Thrp.Time.DL.CAUser.PLMN", "MOCN: CA UE DL time per PLMN", "MOCN throughput"),
        (49, "—", "ChMeas.CQI.CA.PCell / SCell", "Function subset: PCell/SCell CQI", "Measurement of Cell Performance"),
        (50, "—", "ChMeas.MCS.CA.PCell / SCell", "Function subset: PCell/SCell MCS", "Measurement of Cell Performance"),
        (51, "—", "Traffic.MAC.CA.Cell", "Function subset: CA cell MAC", "Measurement of Cell Performance"),
        (52, "—", "DSP UEONLINEINFO / RRC_CONN_RECFG", "sCellToAddModList / sCellToReleaseList; non-zero RB & TBS on PCell and SCell", "Message tracing (MAE-Access)"),
    ]
    for rec in ctrs:
        r = add_ctr(ws, r, *rec)

    # H KPI
    r = section(ws, r, COLS, "H.  KPI formulas for 2CC  (Ch.5.4.3)")
    r = headers(ws, r, ["SN", "KPI", "Formula (from document)", "Unit", "Notes"] + [""] * 5)
    kpis = [
        ("1", "CA UE average DL data rate (last-TTI removed)",
         "(L.Thrp.bits.DL.CAUser − L.Thrp.bits.DL.LastTTI.CAUser) / L.Thrp.Time.DL.RmvLastTTI.CAUser",
         "bit/s", "Compare with non-CA UE rate"),
        ("2", "DL throughput of UEs in 2CC-activated state (only 2CC+3CC enabled example)",
         "(L.Thrp.bits.DL.CAUser − L.Thrp.bits.DL.3CC.CAUser) / (L.Thrp.Time.DL.CAUser − L.Thrp.Time.DL.3CC.CAUser)",
         "bit/s", "Subtracts 3CC (and higher if present) from all-CA. MULTI_CC_STAT_OPT_SW changes whether a UE with 3 of 4 CCs active counts as 3CC."),
        ("3", "General nCC throughput identity",
         "nCC throughput = (CA bits − bits of all CA UEs except nCC) / (CA time − time of all CA UEs except nCC)",
         "bit/s", "Document definition of 2CC-state throughput"),
        ("4", "CA UE E-RAB drop rate",
         "L.E-RAB.AbnormRel.CAUser / (L.E-RAB.AbnormRel.CAUser + L.E-RAB.NormRel.CAUser) × 100%",
         "%", "Service drop"),
        ("5", "CA UE PCell HO success rate",
         "L.HHO.ExecSuccOut.CAUser.PCC / L.HHO.ExecAttOut.CAUser.PCC × 100%",
         "%", ""),
        ("6", "CA UE FDD↔TDD PCell HO success rate",
         "L.HHO.InterFddTdd.ExecSuccOut.CAUser.PCC / L.HHO.InterFddTdd.ExecAttOut.CAUser.PCC × 100%",
         "%", ""),
        ("7", "DL SCell add success rate",
         "L.CA.DLSCell.Add.Succ / L.CA.DLSCell.Add.Att × 100%",
         "%", "Also split Meas vs Blind"),
        ("8", "Measurement-based add success",
         "L.CA.DLSCell.Add.Meas.Succ / L.CA.DLSCell.Add.Meas.Att × 100%",
         "%", "A4 path"),
        ("9", "Blind add success",
         "L.CA.DLSCell.Add.Blind.Succ / L.CA.DLSCell.Add.Blind.Att × 100%",
         "%", "Blind path"),
        ("10", "SCell activation success",
         "L.CA.DLSCell.Act.Succ / L.CA.DLSCell.Act.Att × 100%",
         "%", "MAC CE"),
        ("11", "SCell deactivation success",
         "L.CA.DLSCell.Deact.Succ / L.CA.DLSCell.Deact.Att × 100%",
         "%", "Leaving"),
        ("12", "SCell remove success",
         "L.CA.DLSCell.Rmv.Succ / L.CA.DLSCell.Rmv.Att × 100%",
         "%", "RRC leave"),
        ("13", "E-RAB drop (high-mobility PCC forbid impact)",
         "(L.E-RAB.AbnormRel.eNBTot + L.E-RAB.AbnormRel.HOOut) / L.E-RAB.SuccEst × 100%",
         "%", "Document says this decreases when PCC_ANCHOR_HO_FORBID_SW is ON"),
        ("14", "Contention-based preamble response rate (route penalty)",
         "((L.RA.GrpA.Resp + L.RA.GrpB.Resp) / (L.RA.GrpA.Att + L.RA.GrpB.Att)) × 100%",
         "%", "Referenced by dynamic-route penalty in Ch.4.4"),
    ]
    for i, rec in enumerate(kpis):
        vals = list(rec) + [""] * 5
        r = table_row(ws, r, vals, fills=[alt_fill(i)] * 10, height=40)
        merge(ws, r - 1, 5, r - 1, COLS)

    # I licenses
    r = section(ws, r, COLS, "I.  Licenses — FDD and TDD  (Ch.5.3.1 / 5.3.2)")
    r = headers(ws, r, ["SN", "RAT", "Feature ID", "Feature name", "Model", "Sales unit", "When consumed"] + [""] * 3)
    lic = [
        ("1", "FDD", "LAOFD-001001", "LTE-A Introduction", "LT1SA020CA00", "per cell",
         "Adaptive: every cell on PCC freqs + every SCC-freq cell that participates in DL 2CC. Group: every cell in a group of ≥2 inter-freq FDD cells. Inter-eNB SCells consume the license on the neighbor eNB."),
        ("2", "FDD", "LAOFD-001002", "Carrier Aggregation for Downlink 2CC in 40MHz", "LT1SA040CA00", "per cell",
         "Adaptive: cell already consuming LAOFD-001001 AND CaDl2CCExtSwitch ON. Group: group total BW 20 < BW ≤ 40 MHz, cell already consuming LAOFD-001001."),
        ("3", "TDD", "TDLAOFD-001001", "LTE-A Introduction", "LT1SLTEAID01", "per cell",
         "Every TDD cell in DL 2CC. If aggregated BW ≤ 30 MHz this is the only 2CC license."),
        ("4", "TDD", "TDLAOFD-001002", "Carrier Aggregation for Downlink 2CC in 40MHz", "LT1SC2C40M00", "per cell",
         "Additionally required when aggregated BW is 30 < BW ≤ 40 MHz."),
    ]
    for i, rec in enumerate(lic):
        vals = list(rec) + [""] * 3
        r = table_row(ws, r, vals, fills=[alt_fill(i)] * 10, height=48)
        merge(ws, r - 1, 7, r - 1, COLS)
    r = note_bar(ws, r, COLS, "License insufficiency of certain items blocks CELL ACTIVATION (see License Control Item Lists). Always confirm licenses before ACT CELL.")

    # J extra
    r = section(ws, r, COLS, "J.  Hardware, networking, verification  (Ch.5.3.4–5.3.6, 5.4.2)")
    r = bullets(ws, r, COLS, [
        "Base stations: 3900/5900 macro; DBS3900/DBS5900 LampSite. CA intra- or inter-BBP (inter-BBP BBPs in one BBU).",
        "BBU3900 inter-BBP: LBBPd or LTE UBBP in slot 2 or 3 (ADD BRD) for SRIO (not required for relaxed-BH inter-eNB). BBU3910: UBBPg in slots 0–3.",
        "Intra-band CA RF: one multi-carrier RF, or same-model AAUs, or same-model RRUs, or RFUs with same HW version. Inter-band: no special RF rule.",
        "Need ≥2 frequencies. Intra-band contiguous: center spacing = nominal channel spacing (two 20 MHz → 19.8 MHz; 20+10 → 14.4 MHz). Non-contiguous: spacing ≥ average of the two bandwidths.",
        "Coverage must overlap. Same CP. Candidate SCells = inter-freq neighbors of PCell (except first blind add). Prefer intra-frequency cells on the same BBP if CoMP/SFN also ON.",
        "Verify: Table 5-9 counters non-zero; RRC_CONN_RECFG contains sCellToAddModList; PCell and SCell RB/TBS ≠ 0; DSP UEONLINEINFO.",
        "MAE-Access UE-level: BLER, Channel Quality, Throughput, DL Power Control, MCS Count, User Common Monitoring — start on both PCell and SCell.",
    ])
    ws.row_dimensions[r - 1].height = 120
    return ws
